#!/usr/bin/env python3
"""Latest-quarter source coverage and workbook-fill test.

This is intentionally a test harness, not a final production updater. It copies
the current benchmark workbook, creates clean test sheets, tries to fetch the
latest quarterly rows from machine-readable sources, and fills only values that
were actually extracted.
"""

from __future__ import annotations

import argparse
import io
import json
import math
import re
import shutil
import sys
import time
import urllib.request
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


DEFAULT_WORKBOOK = Path("Бенч финансовой отчетности_мэйджоры.xlsx")
SEC_COMPANY_FACTS_URL = "https://data.sec.gov/api/xbrl/companyfacts/CIK0000042582.json"
SEC_FILING_BASE = "https://www.sec.gov/Archives/edgar/data/42582"


METRICS = [
    "Total Revenues",
    "Cost Of Revenues",
    "Gross Profit",
    "Gross Profit Margin %",
    "Selling General & Admin Expenses",
    "Other Operating Expenses",
    "Operating Income",
    "EBIT Margin %",
    "EBITDA",
    "EBITDA Margin %",
    "Net Income",
    "Accounts Receivable, Total",
    "Inventory",
    "Total Current Liabilities",
    "Total Assets",
    "Total Equity",
    "Total Debt",
    "Capital Expenditure",
    "Levered Free Cash Flow",
    "Cash from Operations",
]


STOCKANALYSIS_REGISTRY = [
    ("Michelin", "EU", "EUR", "https://stockanalysis.com/quote/epa/ML/financials/"),
    ("Continental AG", "EU", "EUR", "https://stockanalysis.com/quote/etr/CON/financials/"),
    ("Pirelli", "EU", "EUR", "https://stockanalysis.com/quote/bit/PIRC/financials/"),
    ("Nokian Tyres", "EU", "EUR", "https://stockanalysis.com/quote/hel/TYRES/financials/"),
    ("Goodyear", "US", "USD", "https://stockanalysis.com/stocks/gt/financials/"),
    ("Bridgestone", "JP", "JPY", "https://stockanalysis.com/quote/tyo/5108/financials/"),
    ("Yokohama", "JP", "JPY", "https://stockanalysis.com/quote/tyo/5101/financials/"),
    ("Sumitomo", "JP", "JPY", "https://stockanalysis.com/quote/tyo/5110/financials/"),
    ("Toyo", "JP", "JPY", "https://stockanalysis.com/quote/tyo/5105/financials/"),
    ("Apollo Tyres", "IN", "INR", "https://stockanalysis.com/quote/nse/APOLLOTYRE/financials/"),
    ("Hankook", "KR", "KRW", "https://stockanalysis.com/quote/krx/161390/financials/"),
    ("Nexen", "KR", "KRW", "https://stockanalysis.com/quote/krx/002350/financials/"),
    ("Kumho", "KR", "KRW", "https://stockanalysis.com/quote/krx/073240/financials/"),
    ("Zhongce Rubber Group", "CN", "CNY", "https://stockanalysis.com/quote/sha/603049/financials/"),
    ("Sailun", "CN", "CNY", "https://stockanalysis.com/quote/sha/601058/financials/"),
    ("Linglong", "CN", "CNY", "https://stockanalysis.com/quote/sha/601966/financials/"),
]


ROW_SYNONYMS = {
    "Total Revenues": [("income", "Revenue"), ("income", "Operating Revenue"), ("income", "Total Revenue")],
    "Cost Of Revenues": [("income", "Cost of Revenue")],
    "Gross Profit": [("income", "Gross Profit")],
    "Selling General & Admin Expenses": [("income", "Selling, General & Admin")],
    "Other Operating Expenses": [("income", "Other Operating Expenses")],
    "Operating Income": [("income", "Operating Income")],
    "EBITDA": [("income", "EBITDA")],
    "Net Income": [("income", "Net Income"), ("income", "Net Income to Company")],
    "Accounts Receivable, Total": [("balance", "Accounts Receivable"), ("balance", "Receivables")],
    "Inventory": [("balance", "Inventory")],
    "Total Current Liabilities": [("balance", "Total Current Liabilities")],
    "Total Assets": [("balance", "Total Assets")],
    "Total Equity": [("balance", "Shareholders' Equity"), ("balance", "Total Common Equity"), ("balance", "Total Equity")],
    "Total Debt": [("balance", "Total Debt")],
    "Capital Expenditure": [("cashflow", "Capital Expenditures")],
    "Levered Free Cash Flow": [("cashflow", "Levered Free Cash Flow"), ("cashflow", "Free Cash Flow")],
    "Cash from Operations": [("cashflow", "Operating Cash Flow")],
}


OFFICIAL_SOURCE_NOTES = {
    "Goodyear": "SEC EDGAR Company Facts is the preferred source and is proven in goodyear_sec_dry_run.py.",
    "Bridgestone": "Official Bridgestone IR publishes Q1 2026 PDFs plus Financial and Sales Data Excel.",
    "Hankook": "OpenDART should be the preferred structured source; StockAnalysis is used here as a readable fallback.",
    "Nexen": "OpenDART should be the preferred structured source; StockAnalysis is used here as a readable fallback.",
    "Kumho": "OpenDART should be the preferred structured source; StockAnalysis is used here as a readable fallback.",
    "Zhongce Rubber Group": "CNINFO / exchange disclosures should be preferred; StockAnalysis is used here as a readable fallback.",
    "Sailun": "CNINFO / exchange disclosures should be preferred; StockAnalysis is used here as a readable fallback.",
    "Linglong": "CNINFO / exchange disclosures should be preferred; StockAnalysis is used here as a readable fallback.",
}


@dataclass
class ExtractedValue:
    company: str
    country: str
    metric: str
    value: float | None
    currency: str
    unit: str
    fiscal_period: str
    period_end: str
    source_type: str
    source_url: str
    source_row: str | None
    status: str
    note: str


def fetch_text(url: str, *, timeout: int = 14, tries: int = 2) -> str:
    last_error: Exception | None = None
    for attempt in range(tries):
        try:
            req = urllib.request.Request(
                url,
                headers={
                    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
                    "Accept-Language": "en-US,en;q=0.9",
                },
            )
            with urllib.request.urlopen(req, timeout=timeout) as response:
                return response.read().decode("utf-8", "ignore")
        except Exception as exc:  # noqa: BLE001 - keep audit note for all fetch errors
            last_error = exc
            time.sleep(0.8 * (attempt + 1))
    raise RuntimeError(f"fetch failed for {url}: {last_error}")


def parse_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text or text in {"-", "—", "–"}:
        return None
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()")
    text = text.replace(",", "").replace("%", "").replace("−", "-").replace("△", "-")
    try:
        number = float(text)
    except ValueError:
        return None
    return -number if negative else number


def table_from_stockanalysis(base_url: str, statement: str) -> tuple[pd.DataFrame, str, str, str]:
    suffix = {
        "income": "",
        "balance": "balance-sheet/",
        "cashflow": "cash-flow-statement/",
    }[statement]
    url = f"{base_url}{suffix}?p=quarterly"
    html = fetch_text(url)
    table = pd.read_html(io.StringIO(html))[0]
    latest_col = table.columns[1]
    if isinstance(latest_col, tuple):
        fiscal_period = str(latest_col[0])
        period_end = str(latest_col[1])
    else:
        fiscal_period = str(latest_col)
        period_end = ""
    return table, fiscal_period, period_end, url


def extract_stockanalysis_company(company: str, country: str, currency: str, base_url: str) -> tuple[list[ExtractedValue], dict[str, Any]]:
    tables: dict[str, pd.DataFrame] = {}
    statement_urls: dict[str, str] = {}
    period_candidates: list[tuple[str, str]] = []
    errors: list[str] = []

    for statement in ("income", "balance", "cashflow"):
        try:
            table, fiscal_period, period_end, url = table_from_stockanalysis(base_url, statement)
            tables[statement] = table
            statement_urls[statement] = url
            period_candidates.append((fiscal_period, period_end))
        except Exception as exc:  # noqa: BLE001 - write to coverage
            errors.append(f"{statement}: {exc}")

    result: list[ExtractedValue] = []
    fiscal_period, period_end = period_candidates[0] if period_candidates else ("", "")

    def lookup(statement: str, labels: list[str]) -> tuple[float | None, str | None]:
        table = tables.get(statement)
        if table is None:
            return None, None
        first_col = table.columns[0]
        latest_col = table.columns[1]
        label_map = {str(row[first_col]).strip(): row[latest_col] for _, row in table.iterrows()}
        for label in labels:
            if label in label_map:
                return parse_number(label_map[label]), label
        return None, None

    raw_values: dict[str, tuple[float | None, str | None, str | None]] = {}
    for metric, synonyms in ROW_SYNONYMS.items():
        value = None
        row_label = None
        source_url = None
        for statement, label in synonyms:
            value, row_label = lookup(statement, [label])
            if value is not None:
                source_url = statement_urls.get(statement)
                break
        raw_values[metric] = (value, row_label, source_url)

    revenue = raw_values.get("Total Revenues", (None, None, None))[0]
    gross_profit = raw_values.get("Gross Profit", (None, None, None))[0]
    operating_income = raw_values.get("Operating Income", (None, None, None))[0]
    ebitda = raw_values.get("EBITDA", (None, None, None))[0]
    if revenue:
        raw_values["Gross Profit Margin %"] = (
            round(gross_profit / revenue, 6) if gross_profit is not None else None,
            "computed: Gross Profit / Revenue",
            statement_urls.get("income"),
        )
        raw_values["EBIT Margin %"] = (
            round(operating_income / revenue, 6) if operating_income is not None else None,
            "computed: Operating Income / Revenue",
            statement_urls.get("income"),
        )
        raw_values["EBITDA Margin %"] = (
            round(ebitda / revenue, 6) if ebitda is not None else None,
            "computed: EBITDA / Revenue",
            statement_urls.get("income"),
        )

    latest_is_quarter = fiscal_period.upper().startswith("Q")
    for metric in METRICS:
        value, row_label, source_url = raw_values.get(metric, (None, None, None))
        result.append(
            ExtractedValue(
                company=company,
                country=country,
                metric=metric,
                value=value,
                currency=currency,
                unit="millions" if not metric.endswith("%") else "ratio",
                fiscal_period=fiscal_period,
                period_end=period_end,
                source_type="aggregator_readable_html",
                source_url=source_url or base_url,
                source_row=row_label,
                status="filled" if value is not None and latest_is_quarter else ("stale_period" if value is not None else "missing"),
                note="" if latest_is_quarter else "StockAnalysis latest periodic table is not a Q-period.",
            )
        )

    coverage = {
        "company": company,
        "source_type": "aggregator_readable_html",
        "base_url": base_url,
        "fiscal_period": fiscal_period,
        "period_end": period_end,
        "statement_tables": sorted(tables),
        "errors": errors,
        "filled_metrics": sum(1 for row in result if row.value is not None and row.status == "filled"),
        "available_metrics": sum(1 for row in result if row.value is not None),
        "official_note": OFFICIAL_SOURCE_NOTES.get(company, ""),
    }
    return result, coverage


def write_output_workbook(source_path: Path, output_path: Path, values: list[ExtractedValue], coverage_rows: list[dict[str, Any]]) -> None:
    shutil.copy2(source_path, output_path)
    wb = load_workbook(output_path)
    for sheet_name in ("Latest_Q_Fill_Test", "Latest_Q_Evidence", "Source_Coverage"):
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    fill = wb.create_sheet("Latest_Q_Fill_Test", 0)
    evidence = wb.create_sheet("Latest_Q_Evidence", 1)
    coverage = wb.create_sheet("Source_Coverage", 2)

    companies = []
    seen = set()
    for row in values:
        if row.company not in seen:
            companies.append(row.company)
            seen.add(row.company)

    value_map = {(row.company, row.metric): row for row in values}
    fill_headers = ["company", "country", "fiscal_period", "period_end", "currency"] + METRICS
    fill.append(fill_headers)
    for company in companies:
        company_values = [row for row in values if row.company == company]
        first = company_values[0]
        output_row = [company, first.country, first.fiscal_period, first.period_end, first.currency]
        for metric in METRICS:
            item = value_map.get((company, metric))
            output_row.append(item.value if item and item.status == "filled" else None)
        fill.append(output_row)

    evidence_headers = [
        "company",
        "country",
        "metric",
        "value",
        "currency",
        "unit",
        "fiscal_period",
        "period_end",
        "status",
        "source_type",
        "source_row",
        "source_url",
        "note",
    ]
    evidence.append(evidence_headers)
    for row in values:
        evidence.append(
            [
                row.company,
                row.country,
                row.metric,
                row.value,
                row.currency,
                row.unit,
                row.fiscal_period,
                row.period_end,
                row.status,
                row.source_type,
                row.source_row,
                row.source_url,
                row.note,
            ]
        )

    coverage_headers = [
        "company",
        "source_type",
        "fiscal_period",
        "period_end",
        "filled_metrics",
        "available_metrics",
        "statement_tables",
        "base_url",
        "errors",
        "official_note",
    ]
    coverage.append(coverage_headers)
    for row in coverage_rows:
        coverage.append(
            [
                row.get("company"),
                row.get("source_type"),
                row.get("fiscal_period"),
                row.get("period_end"),
                row.get("filled_metrics"),
                row.get("available_metrics"),
                ", ".join(row.get("statement_tables", [])),
                row.get("base_url"),
                " | ".join(row.get("errors", [])),
                row.get("official_note"),
            ]
        )

    for ws in (fill, evidence, coverage):
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = PatternFill(fill_type="solid", fgColor="1F4E78")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        widths = {
            "A": 24,
            "B": 10,
            "C": 14,
            "D": 24,
            "E": 10,
            "F": 14,
            "G": 14,
            "H": 14,
            "I": 16,
            "J": 18,
            "K": 18,
            "L": 16,
            "M": 18,
            "N": 18,
            "O": 18,
            "P": 18,
            "Q": 18,
            "R": 18,
            "S": 18,
            "T": 18,
            "U": 18,
            "V": 18,
            "W": 18,
            "X": 18,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

    for row in evidence.iter_rows(min_row=2, min_col=12, max_col=12):
        cell = row[0]
        if cell.value and str(cell.value).startswith("http"):
            cell.hyperlink = cell.value
            cell.style = "Hyperlink"
    for row in coverage.iter_rows(min_row=2, min_col=8, max_col=8):
        cell = row[0]
        if cell.value and str(cell.value).startswith("http"):
            cell.hyperlink = cell.value
            cell.style = "Hyperlink"

    wb.save(output_path)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--output-dir", type=Path, default=Path("outputs"))
    args = parser.parse_args()

    source_path = args.workbook.resolve()
    if not source_path.exists():
        print(f"Workbook not found: {source_path}", file=sys.stderr)
        return 2

    run_id = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    out_dir = (args.output_dir / f"latest_quarter_fill_test_{run_id}").resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    values: list[ExtractedValue] = []
    coverage_rows: list[dict[str, Any]] = []
    for company, country, currency, base_url in STOCKANALYSIS_REGISTRY:
        try:
            extracted, coverage = extract_stockanalysis_company(company, country, currency, base_url)
        except Exception as exc:  # noqa: BLE001 - record source failure
            extracted = [
                ExtractedValue(
                    company=company,
                    country=country,
                    metric=metric,
                    value=None,
                    currency=currency,
                    unit="millions" if not metric.endswith("%") else "ratio",
                    fiscal_period="",
                    period_end="",
                    source_type="aggregator_readable_html",
                    source_url=base_url,
                    source_row=None,
                    status="source_failed",
                    note=str(exc),
                )
                for metric in METRICS
            ]
            coverage = {
                "company": company,
                "source_type": "aggregator_readable_html",
                "base_url": base_url,
                "fiscal_period": "",
                "period_end": "",
                "statement_tables": [],
                "errors": [str(exc)],
                "filled_metrics": 0,
                "available_metrics": 0,
                "official_note": OFFICIAL_SOURCE_NOTES.get(company, ""),
            }
        values.extend(extracted)
        coverage_rows.append(coverage)
        print(
            f"{company}: {coverage['fiscal_period']} {coverage['period_end']} "
            f"filled={coverage['filled_metrics']} available={coverage['available_metrics']} errors={len(coverage['errors'])}",
            flush=True,
        )

    workbook_output = out_dir / f"{source_path.stem}_latest_quarter_fill_test.xlsx"
    json_output = out_dir / "latest_quarter_fill_test.json"
    write_output_workbook(source_path, workbook_output, values, coverage_rows)
    json_output.write_text(
        json.dumps(
            {
                "run_id": run_id,
                "note": "StockAnalysis is used here as a readable aggregator test source, not as final source-of-truth.",
                "metrics": METRICS,
                "coverage": coverage_rows,
                "values": [asdict(row) for row in values],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    print(f"Wrote {workbook_output}")
    print(f"Wrote {json_output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
