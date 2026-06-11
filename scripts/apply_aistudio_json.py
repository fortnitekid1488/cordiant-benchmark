#!/usr/bin/env python3
"""Apply LLM JSON extraction output to a copied benchmark workbook."""

from __future__ import annotations

import argparse
import copy
import json
import re
import sys
import warnings
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from workbook_paths import DEFAULT_WORKBOOK, resolve_workbook_path


ROOT = Path(__file__).resolve().parents[1]
SUMMARY_SHEET = "Свод"
OUTPUT_MAX_ROW = 59
OUTPUT_MAX_COL = 31
MIN_VISIBLE_COMPANY_COL_WIDTH = 11.5
EXCEL_ERROR_CODES = {"#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!", "#REF!", "#VALUE!"}
PERCENT_ROWS = {9, 15, 17, 25, 27, 38, 41, 43, 46, 49, 57, 59}
DEFAULT_LC_USD_RATES = {
    "eu": 1.1306,
    "us": 1.0,
    "jp": 149.65,
    "kr": 1421.6,
    "cn": 7.1889,
}
QUARTER_PERIOD_RE = re.compile(
    r"\b(q[1-4]|[1-4]q)\b|first\s+quarter|second\s+quarter|third\s+quarter|fourth\s+quarter|"
    r"\bquarterly\b|three\s+months|3\s+months",
    re.I,
)
CUMULATIVE_PERIOD_RE = re.compile(
    r"\b(h1|1h|h2|2h|9m|6m|ytd|ttm)\b|half[-\s]?year|six\s+months|nine\s+months",
    re.I,
)
INTERIM_PERIOD_RE = re.compile(r"\binterim\b", re.I)
ANNUAL_PERIOD_RE = re.compile(r"\b(fy|full[-\s]?year|annual)\b|year\s+ended|12\s+months|twelve\s+months", re.I)
METRICS = [
    "Total Revenues",
    "# employees",
    "Cost Of Revenues",
    "Gross Profit",
    "Gross Profit Margin %",
    "Other Operating Expenses, Total",
    "R&D Expenses",
    "Selling General & Admin Expenses",
    "Other Operating Expenses",
    "Operating Income",
    "EBIT Margin %",
    "EBITDA",
    "EBITDA Margin %",
    "Total Receivables",
    "Net Income",
    "Accounts Receivable, Total",
    "Inventory",
    "Total Current Liabilities",
    "Total Assets",
    "Total Equity",
    "Total Debt",
    "Capital Expenditure",
    "Levered Free Cash Flow",
    "Cash from Operations",
]

TOP_METRIC_ROWS = {
    "Total Revenues": 4,
    "# employees": 5,
    "Cost Of Revenues": 7,
    "Gross Profit": 8,
    "Gross Profit Margin %": 9,
    "Other Operating Expenses, Total": 10,
    "R&D Expenses": 11,
    "Selling General & Admin Expenses": 12,
    "Other Operating Expenses": 13,
    "Operating Income": 14,
    "EBIT Margin %": 15,
    "EBITDA": 16,
    "EBITDA Margin %": 17,
    "Total Receivables": 18,
    "Accounts Receivable, Total": 19,
    "Inventory": 21,
    "Total Debt": 23,
    "Capital Expenditure": 24,
    "Levered Free Cash Flow": 26,
}

COMPANY_ALIASES = {
    "continental": "continental ag",
    "nokian": "nokian tyres",
    "goodyear": "goodyear",
    "bridgestone": "bridgestone",
    "yokohama": "yokohama",
    "sumitomo": "sumitomo",
    "toyo": "toyo",
    "hankook": "hankook",
    "nexen": "nexen",
    "kumho": "kumho",
    "zhongce": "zhongce rubber group",
    "sailun": "sailun",
    "linglong": "linglong",
    "michelin": "michelin",
    "pirelli": "pirelli",
}


def load_json_files(paths: list[Path]) -> list[dict[str, Any]]:
    payloads = []
    for path in paths:
        if not path.exists():
            continue
        try:
            payloads.append(json.loads(path.read_text(encoding="utf-8")))
        except json.JSONDecodeError as exc:
            raise ValueError(f"Invalid JSON in {path}: {exc}") from exc
    return payloads


def collect_companies(payloads: list[dict[str, Any]]) -> list[dict[str, Any]]:
    companies = []
    for payload in payloads:
        if isinstance(payload.get("companies"), list):
            companies.extend(payload["companies"])
        elif isinstance(payload, list):
            companies.extend(payload)
        else:
            raise ValueError("LLM JSON must contain a `companies` array.")
    return companies


def fact_map(company: dict[str, Any]) -> dict[str, dict[str, Any]]:
    result = {}
    for fact in company.get("facts", []):
        metric = fact.get("metric")
        if metric:
            result[metric] = fact
    return result


def normalize_label(value: Any) -> str:
    return " ".join(str(value or "").replace("\xa0", " ").strip().lower().split())


def normalize_company(value: Any) -> str:
    text = normalize_label(value)
    for suffix in (" tyresa",):
        if text.endswith(suffix):
            text = text[: -len(suffix)].strip()
    return text


def company_keys(company: dict[str, Any]) -> list[str]:
    candidates = [
        company.get("company_key"),
        company.get("company_name"),
    ]
    key = normalize_company(company.get("company_key"))
    if key in COMPANY_ALIASES:
        candidates.append(COMPANY_ALIASES[key])
    return [normalize_company(candidate) for candidate in candidates if normalize_company(candidate)]


def is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def usable_fact(fact: dict[str, Any] | None) -> bool:
    if not fact or fact.get("review_required"):
        return False
    if not is_number(fact.get("value")):
        return False
    has_source = bool(fact.get("source_file") or fact.get("source_url"))
    has_evidence = bool(fact.get("page") or fact.get("table_label") or fact.get("evidence_quote"))
    return has_source and has_evidence


def parse_period_date(value: Any):
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value).strip()).date()
    except ValueError:
        return None


def period_text(company: dict[str, Any]) -> str:
    return " ".join(
        str(value or "")
        for value in (
            company.get("period_label"),
            company.get("period_start"),
            company.get("period_end"),
        )
    )


def period_duration_days(company: dict[str, Any]) -> int | None:
    start = parse_period_date(company.get("period_start"))
    end = parse_period_date(company.get("period_end"))
    if not start or not end:
        return None
    return (end - start).days + 1


def period_validation_issue(company: dict[str, Any], mode: str) -> str | None:
    text = period_text(company)
    label = str(company.get("period_label") or "").strip()
    end = parse_period_date(company.get("period_end"))
    duration = period_duration_days(company)
    has_quarter = bool(QUARTER_PERIOD_RE.search(text))
    has_cumulative = bool(CUMULATIVE_PERIOD_RE.search(text))
    has_annual = bool(ANNUAL_PERIOD_RE.search(text))

    if not label and not end and duration is None:
        return "Missing period_label/period_end; cannot verify selected reporting period."

    if mode == "annual":
        if has_quarter or has_cumulative or INTERIM_PERIOD_RE.search(text):
            return "Annual mode received quarterly/interim/cumulative period metadata."
        if duration is not None and duration < 300:
            return f"Annual mode expected a full fiscal year, got {duration} days."
        if not label and duration is None:
            return "Annual mode requires a period label or start/end dates."
        if not (has_annual or re.search(r"\b20\d{2}\b", label) or duration is not None or end is not None):
            return "Annual mode period does not identify a fiscal year."
        return None

    if mode == "quarterly":
        if has_cumulative:
            return "Quarterly mode received cumulative H1/9M/YTD/TTM period metadata."
        if has_annual and not has_quarter:
            return "Quarterly mode received annual/full-year period metadata."
        if duration is not None and not 70 <= duration <= 115:
            return f"Quarterly mode expected a standalone quarter, got {duration} days."
        if not has_quarter and duration is None:
            return "Quarterly mode period does not identify a standalone quarter."
        return None

    return None


def period_freshness_warning(company: dict[str, Any], mode: str, now: datetime | None = None) -> str | None:
    end = parse_period_date(company.get("period_end"))
    if not end:
        return None
    today = (now or datetime.now(timezone.utc)).date()
    if mode == "annual" and end.year < today.year - 1:
        return f"Annual period_end {end.isoformat()} is older than the prior fiscal year for this run date."
    if mode == "quarterly" and (today - end).days > 220:
        return f"Quarterly period_end {end.isoformat()} is more than 220 days before this run date."
    return None


def safe_div(numerator: Any, denominator: Any) -> float | None:
    if not is_number(numerator) or not is_number(denominator) or denominator == 0:
        return None
    return float(numerator) / float(denominator)


def cell_number(ws, row: int, col: int) -> float | None:
    value = ws.cell(row, col).value
    return float(value) if is_number(value) else None


def exchange_rate_for_column(ws, col: int) -> float | None:
    rate = cell_number(ws, 33, col)
    if rate is not None and rate != 0:
        return rate
    country = normalize_label(ws.cell(32, col).value)
    rate = DEFAULT_LC_USD_RATES.get(country)
    if rate is not None:
        ws.cell(33, col).value = rate
    return rate


def set_if_number(ws, row: int, col: int, value: float | None, fallback: Any = None) -> None:
    ws.cell(row, col).value = value if value is not None else fallback


def copy_summary_sheet(workbook_path: Path):
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        style_wb = load_workbook(workbook_path, data_only=False, keep_links=False)
        value_wb = load_workbook(workbook_path, data_only=True, keep_links=False)

    source = style_wb[SUMMARY_SHEET]
    values = value_wb[SUMMARY_SHEET]
    output_wb = Workbook()
    ws = output_wb.active
    ws.title = SUMMARY_SHEET
    ws.sheet_view.showGridLines = source.sheet_view.showGridLines
    ws.freeze_panes = "D32"
    ws.auto_filter.ref = f"A31:{get_column_letter(OUTPUT_MAX_COL)}{OUTPUT_MAX_ROW}"

    for col in range(1, OUTPUT_MAX_COL + 1):
        letter = get_column_letter(col)
        src_dim = source.column_dimensions[letter]
        dst_dim = ws.column_dimensions[letter]
        dst_dim.width = src_dim.width
        if 4 <= col <= 18:
            dst_dim.width = max(dst_dim.width or 0, MIN_VISIBLE_COMPANY_COL_WIDTH)
        dst_dim.hidden = src_dim.hidden
        dst_dim.outlineLevel = src_dim.outlineLevel

    for row in range(1, OUTPUT_MAX_ROW + 1):
        src_dim = source.row_dimensions[row]
        dst_dim = ws.row_dimensions[row]
        dst_dim.height = src_dim.height
        dst_dim.hidden = src_dim.hidden
        dst_dim.outlineLevel = src_dim.outlineLevel
        for col in range(1, OUTPUT_MAX_COL + 1):
            src_cell = source.cell(row, col)
            value_cell = values.cell(row, col)
            dst_cell = ws.cell(row, col)
            dst_cell.value = "N/A" if value_cell.value in EXCEL_ERROR_CODES else value_cell.value
            dst_cell.font = copy.copy(src_cell.font)
            dst_cell.fill = copy.copy(src_cell.fill)
            dst_cell.border = copy.copy(src_cell.border)
            dst_cell.alignment = copy.copy(src_cell.alignment)
            dst_cell.protection = copy.copy(src_cell.protection)
            dst_cell.number_format = src_cell.number_format
            if src_cell.hyperlink:
                dst_cell._hyperlink = copy.copy(src_cell.hyperlink)

    for row in PERCENT_ROWS:
        for col in range(4, 19):
            ws.cell(row, col).number_format = "0%"

    return output_wb, ws


def build_company_column_map(ws) -> dict[str, int]:
    result: dict[str, int] = {}
    for col in range(4, 19):
        for row in (1, 31):
            normalized = normalize_company(ws.cell(row, col).value)
            if normalized:
                result[normalized] = col
    for alias, target in COMPANY_ALIASES.items():
        if target in result:
            result[alias] = result[target]
    return result


def provenance_comment(company: dict[str, Any], metric: str, fact: dict[str, Any]) -> str:
    payload = {
        "company_key": company.get("company_key"),
        "company_name": company.get("company_name"),
        "metric": metric,
        "period_label": company.get("period_label"),
        "period_start": company.get("period_start"),
        "period_end": company.get("period_end"),
        "currency": company.get("currency"),
        "unit": company.get("unit"),
        "source_file": fact.get("source_file"),
        "source_url": fact.get("source_url"),
        "source_type": fact.get("source_type"),
        "page": fact.get("page"),
        "table_label": fact.get("table_label"),
        "confidence": fact.get("confidence"),
        "evidence_quote": fact.get("evidence_quote"),
    }
    return json.dumps(payload, ensure_ascii=False, indent=2)


def converted_to_usd(ws, row: int, col: int) -> float | None:
    local_value = cell_number(ws, row, col)
    if local_value is None:
        return None
    country = normalize_label(ws.cell(32, col).value)
    if country == "us":
        exchange_rate_for_column(ws, col)
        return local_value
    rate = exchange_rate_for_column(ws, col)
    if rate is None or rate == 0:
        return None
    if country == "eu":
        return local_value * rate
    return local_value / rate


def recalc_company_column(ws, col: int) -> None:
    revenue = cell_number(ws, 4, col)
    employees = cell_number(ws, 5, col)
    gross_profit = cell_number(ws, 8, col)
    operating_income = cell_number(ws, 14, col)
    ebitda = cell_number(ws, 16, col)
    receivables = cell_number(ws, 19, col)
    inventory = cell_number(ws, 21, col)
    capex = cell_number(ws, 24, col)
    fcf = cell_number(ws, 26, col)

    set_if_number(ws, 6, col, safe_div(revenue, employees), ws.cell(6, col).value)
    set_if_number(ws, 9, col, safe_div(gross_profit, revenue), ws.cell(9, col).value)
    set_if_number(ws, 15, col, safe_div(operating_income, revenue), ws.cell(15, col).value)
    set_if_number(ws, 17, col, safe_div(ebitda, revenue), ws.cell(17, col).value)
    set_if_number(ws, 20, col, safe_div(revenue, receivables), ws.cell(20, col).value)
    set_if_number(ws, 22, col, safe_div(revenue, inventory), ws.cell(22, col).value)
    set_if_number(ws, 25, col, safe_div(capex, revenue), ws.cell(25, col).value)
    set_if_number(ws, 27, col, safe_div(fcf, ebitda), ws.cell(27, col).value)

    bottom_updates = {
        34: converted_to_usd(ws, 4, col),
        36: converted_to_usd(ws, 7, col),
        37: converted_to_usd(ws, 8, col),
        39: converted_to_usd(ws, 10, col),
        40: converted_to_usd(ws, 11, col),
        42: converted_to_usd(ws, 12, col),
        44: converted_to_usd(ws, 13, col),
        45: converted_to_usd(ws, 14, col),
        47: converted_to_usd(ws, 16, col),
        50: converted_to_usd(ws, 18, col),
        51: converted_to_usd(ws, 19, col),
        53: converted_to_usd(ws, 21, col),
        55: converted_to_usd(ws, 23, col),
        56: converted_to_usd(ws, 24, col),
        58: converted_to_usd(ws, 26, col),
    }
    for row, value in bottom_updates.items():
        if value is not None:
            ws.cell(row, col).value = value

    revenue_usd = cell_number(ws, 34, col)
    gross_profit_usd = cell_number(ws, 37, col)
    research_development_usd = cell_number(ws, 40, col)
    sga = cell_number(ws, 12, col)
    operating_income_usd = cell_number(ws, 45, col)
    ebitda_usd = cell_number(ws, 47, col)
    receivables_usd = cell_number(ws, 51, col)
    inventory_usd = cell_number(ws, 53, col)
    capex_usd = cell_number(ws, 56, col)
    fcf_usd = cell_number(ws, 58, col)

    set_if_number(ws, 35, col, safe_div(revenue_usd * 1000 if revenue_usd is not None else None, employees), ws.cell(35, col).value)
    set_if_number(ws, 38, col, safe_div(gross_profit_usd, revenue_usd), ws.cell(38, col).value)
    set_if_number(ws, 41, col, safe_div(research_development_usd, revenue_usd), ws.cell(41, col).value)
    set_if_number(ws, 43, col, safe_div(sga, revenue), ws.cell(43, col).value)
    set_if_number(ws, 46, col, safe_div(operating_income_usd, revenue_usd), ws.cell(46, col).value)
    set_if_number(ws, 48, col, safe_div(ebitda_usd * 1000 if ebitda_usd is not None else None, employees), ws.cell(48, col).value)
    set_if_number(ws, 49, col, safe_div(ebitda_usd, revenue_usd), ws.cell(49, col).value)
    set_if_number(
        ws,
        52,
        col,
        safe_div(365 * receivables_usd if receivables_usd is not None else None, revenue_usd),
        ws.cell(52, col).value,
    )
    set_if_number(
        ws,
        54,
        col,
        safe_div(365 * inventory_usd if inventory_usd is not None else None, revenue_usd),
        ws.cell(54, col).value,
    )
    set_if_number(
        ws,
        57,
        col,
        safe_div(abs(capex_usd) if capex_usd is not None else None, revenue_usd),
        ws.cell(57, col).value,
    )
    set_if_number(ws, 59, col, safe_div(fcf_usd, ebitda_usd), ws.cell(59, col).value)


def apply_companies_to_summary(ws, companies: list[dict[str, Any]], mode: str) -> dict[str, Any]:
    column_map = build_company_column_map(ws)
    audit: dict[str, Any] = {"applied": [], "review": [], "skipped_companies": [], "period_warnings": []}

    for company in companies:
        col = next((column_map[key] for key in company_keys(company) if key in column_map), None)
        if col is None:
            audit["skipped_companies"].append(
                {
                    "company_key": company.get("company_key"),
                    "company_name": company.get("company_name"),
                    "reason": "No matching visible company column in the source `Свод` table.",
                }
            )
            continue

        period_issue = period_validation_issue(company, mode)
        if period_issue:
            audit["skipped_companies"].append(
                {
                    "company_key": company.get("company_key"),
                    "company_name": company.get("company_name"),
                    "period_label": company.get("period_label"),
                    "period_start": company.get("period_start"),
                    "period_end": company.get("period_end"),
                    "reason": period_issue,
                }
            )
            continue

        freshness_warning = period_freshness_warning(company, mode)
        if freshness_warning:
            audit["period_warnings"].append(
                {
                    "company_key": company.get("company_key"),
                    "company_name": company.get("company_name"),
                    "period_label": company.get("period_label"),
                    "period_start": company.get("period_start"),
                    "period_end": company.get("period_end"),
                    "warning": freshness_warning,
                }
            )

        facts = fact_map(company)
        usable_metrics = {
            metric: facts.get(metric)
            for metric in TOP_METRIC_ROWS
            if usable_fact(facts.get(metric))
        }
        if not usable_metrics:
            audit["skipped_companies"].append(
                {
                    "company_key": company.get("company_key"),
                    "company_name": company.get("company_name"),
                    "reason": "No non-review values with source evidence for visible summary metrics.",
                }
            )

        for metric, row in TOP_METRIC_ROWS.items():
            fact = facts.get(metric)
            if metric in usable_metrics:
                if company.get("period_label"):
                    ws.cell(3, col).value = company["period_label"]
                if company.get("currency"):
                    ws.cell(28, col).value = f"* In Millions of {company['currency']}"
                cell = ws.cell(row, col)
                cell.value = fact["value"]
                cell.comment = Comment(provenance_comment(company, metric, fact), "Codex")
                audit["applied"].append(
                    {
                        "company_key": company.get("company_key"),
                        "company_name": company.get("company_name"),
                        "metric": metric,
                        "cell": cell.coordinate,
                        "value": fact["value"],
                        "source_file": fact.get("source_file"),
                        "source_url": fact.get("source_url"),
                        "confidence": fact.get("confidence"),
                    }
                )
            elif fact:
                audit["review"].append(
                    {
                        "company_key": company.get("company_key"),
                        "company_name": company.get("company_name"),
                        "metric": metric,
                        "value": fact.get("value"),
                        "reason": fact.get("review_reason") or "Fact is marked review_required or lacks source evidence.",
                    }
                )

        if usable_metrics:
            recalc_company_column(ws, col)

    return audit


def setup_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for i, width in enumerate([24, 14, 14, 14, 10] + [18] * 25, start=1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "Z"].width = width


def load_workbook_pair(path: Path):
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        formulas = load_workbook(path, keep_links=False)
        cached_values = load_workbook(path, data_only=True, keep_links=False)
    return formulas, cached_values


def sanitize_for_excel_open(wb, cached_wb) -> dict[str, int]:
    stats = {
        "external_links_removed": len(getattr(wb, "_external_links", [])),
        "external_formulas_replaced": 0,
        "drawings_removed": 0,
        "defined_names_removed": len(wb.defined_names),
    }
    wb._external_links = []
    wb.defined_names.clear()

    for ws in wb.worksheets:
        cached_ws = cached_wb[ws.title] if cached_wb and ws.title in cached_wb.sheetnames else None
        images = getattr(ws, "_images", [])
        charts = getattr(ws, "_charts", [])
        if images:
            stats["drawings_removed"] += len(images)
            ws._images = []
        if charts:
            stats["drawings_removed"] += len(charts)
            ws._charts = []
        if getattr(ws, "_drawing", None) is not None:
            ws._drawing = None

        if cached_ws is None:
            continue
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("=") and "[" in cell.value:
                    cell.value = cached_ws[cell.coordinate].value
                    stats["external_formulas_replaced"] += 1
    return stats


def set_initial_sheet(wb, sheet_name: str) -> None:
    wb.active = wb.sheetnames.index(sheet_name)
    for ws in wb.worksheets:
        ws.sheet_view.tabSelected = ws.title == sheet_name


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--json-dir", type=Path)
    parser.add_argument("--output-dir", type=Path, default=ROOT / "outputs")
    parser.add_argument("--mode", choices=["quarterly", "annual"], default="quarterly")
    args = parser.parse_args()

    try:
        args.workbook = resolve_workbook_path(args.workbook)
    except FileNotFoundError as exc:
        print(str(exc), file=sys.stderr)
        return 2

    json_dir = args.json_dir
    if json_dir is None:
        latest_path_file = ROOT / "outputs" / "latest_aistudio_package_path.txt"
        mode_path_file = ROOT / "outputs" / f"latest_aistudio_{args.mode}_package_path.txt"
        latest_path_file = mode_path_file if mode_path_file.exists() else latest_path_file
        if latest_path_file.exists():
            json_dir = Path(latest_path_file.read_text(encoding="utf-8").strip()) / "aistudio_json"
        else:
            print("No --json-dir provided and latest package path not found.", file=sys.stderr)
            return 2

    json_files = sorted(json_dir.glob("*.json"))
    if not json_files:
        print(f"No JSON files found in {json_dir}", file=sys.stderr)
        return 2

    payloads = load_json_files(json_files)
    companies = collect_companies(payloads)

    run_id = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    out_dir = (args.output_dir / f"aistudio_{args.mode}_excel_update_{run_id}").resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    suffix = "aistudio_annual_update" if args.mode == "annual" else "aistudio_quarterly_update"
    output_workbook = out_dir / f"{args.workbook.stem}_{suffix}_{run_id}.xlsx"

    wb, ws = copy_summary_sheet(args.workbook)
    apply_audit = apply_companies_to_summary(ws, companies, args.mode)
    audit = {
        "created_at": datetime.now(timezone.utc).isoformat(),
        "mode": args.mode,
        "source_workbook": str(args.workbook.resolve()),
        "json_files": [str(path.resolve()) for path in json_files],
        "output_sheet": SUMMARY_SHEET,
        "output_range": f"A1:{get_column_letter(OUTPUT_MAX_COL)}{OUTPUT_MAX_ROW}",
        "one_sheet_workbook": True,
        "companies_in_json": len(companies),
        **apply_audit,
    }
    set_initial_sheet(wb, SUMMARY_SHEET)
    wb.save(output_workbook)
    audit_path = out_dir / f"{output_workbook.stem}_provenance.json"
    audit_path.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    print(output_workbook)
    print(audit_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
