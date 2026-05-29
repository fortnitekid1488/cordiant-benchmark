#!/usr/bin/env python3
"""Dry-run update proof for one structured filer in the tire benchmark workbook.

The script fetches Goodyear facts from the SEC Company Facts API, maps a small
set of workbook metrics to XBRL tags, compares them with the current workbook,
and writes an auditable JSON plus a copied workbook containing an
Automation_POC sheet. It intentionally does not modify the source workbook.
"""

from __future__ import annotations

import argparse
import json
import shutil
import sys
import urllib.request
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


DEFAULT_WORKBOOK = Path("Бенч финансовой отчетности_мэйджоры.xlsx")
SEC_COMPANY_FACTS_URL = "https://data.sec.gov/api/xbrl/companyfacts/CIK0000042582.json"
SEC_FILING_BASE = "https://www.sec.gov/Archives/edgar/data/42582"


METRIC_MAP = {
    "Total Revenues": {
        "tag": "RevenueFromContractWithCustomerExcludingAssessedTax",
        "unit": "USD",
        "statement": "income_statement",
        "duration": True,
        "workbook_label": "Total Revenues",
    },
    "Net Income": {
        "tag": "NetIncomeLoss",
        "unit": "USD",
        "statement": "income_statement",
        "duration": True,
        "workbook_label": "Net Income",
    },
    "Total Assets": {
        "tag": "Assets",
        "unit": "USD",
        "statement": "balance_sheet",
        "duration": False,
        "workbook_label": "Total Assets",
    },
    "Total Current Liabilities": {
        "tag": "LiabilitiesCurrent",
        "unit": "USD",
        "statement": "balance_sheet",
        "duration": False,
        "workbook_label": "Total Current Liabilities",
    },
    "Total Equity": {
        "tag": "StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest",
        "unit": "USD",
        "statement": "balance_sheet",
        "duration": False,
        "workbook_label": "Total Equity",
    },
    "Operating Cash Flow": {
        "tag": "NetCashProvidedByUsedInOperatingActivities",
        "unit": "USD",
        "statement": "cash_flow",
        "duration": True,
        "workbook_label": "Cash from Operations",
    },
    "Capital Expenditure": {
        "tag": "PaymentsToAcquirePropertyPlantAndEquipment",
        "unit": "USD",
        "statement": "cash_flow",
        "duration": True,
        "sign": -1,
        "workbook_label": "Capital Expenditure",
    },
}


@dataclass
class ExtractedFact:
    metric: str
    xbrl_tag: str
    value_usd: int | float
    value_musd: float
    period: str
    fiscal_year: int
    fiscal_period: str
    form: str
    filed: str
    accession: str
    source_url: str
    current_workbook_value_musd: Any
    delta_musd: float | None
    status: str


def fetch_json(url: str) -> dict[str, Any]:
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": "Cordiant financial benchmark automation proof nikita@example.local",
            "Accept": "application/json",
        },
    )
    with urllib.request.urlopen(request, timeout=60) as response:
        return json.load(response)


def filing_url(accession: str) -> str:
    accession_no_dashes = accession.replace("-", "")
    return f"{SEC_FILING_BASE}/{accession_no_dashes}/{accession}-index.html"


def fact_candidates(company_facts: dict[str, Any], tag: str, unit: str) -> list[dict[str, Any]]:
    try:
        return company_facts["facts"]["us-gaap"][tag]["units"][unit]
    except KeyError:
        return []


def pick_latest_fact(
    facts: list[dict[str, Any]],
    *,
    target_form: str,
    fiscal_periods: tuple[str, ...],
    duration: bool,
) -> dict[str, Any] | None:
    candidates = [
        fact
        for fact in facts
        if fact.get("form") == target_form
        and fact.get("fp") in fiscal_periods
        and fact.get("fy") is not None
        and fact.get("end")
    ]
    if duration:
        candidates = [fact for fact in candidates if fact.get("start")]
    return max(
        candidates,
        key=lambda fact: (
            fact.get("end", ""),
            fact.get("filed", ""),
            fact.get("start", ""),
        ),
        default=None,
    )


def find_current_value(ws, label: str, preferred_year: int = 2025) -> Any:
    """Find a metric in the existing Goodyear sheet and return the value for year."""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == label:
                header_row = None
                for probe in range(cell.row, max(cell.row - 24, 0), -1):
                    row_values = [ws.cell(probe, col).value for col in range(1, ws.max_column + 1)]
                    if "Period Ending:" in row_values:
                        header_row = probe
                        break
                if not header_row:
                    continue
                for col in range(1, ws.max_column + 1):
                    if ws.cell(header_row, col).value == preferred_year:
                        return ws.cell(cell.row, col).value
    return None


def build_rows(company_facts: dict[str, Any], workbook_path: Path) -> list[ExtractedFact]:
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    ws = wb["GoodYear"]

    rows: list[ExtractedFact] = []
    for metric, config in METRIC_MAP.items():
        facts = fact_candidates(company_facts, config["tag"], config["unit"])
        latest_annual = pick_latest_fact(
            facts,
            target_form="10-K",
            fiscal_periods=("FY",),
            duration=config["duration"],
        )
        latest_quarter = pick_latest_fact(
            facts,
            target_form="10-Q",
            fiscal_periods=("Q1", "Q2", "Q3"),
            duration=config["duration"],
        )

        # Annual facts are comparable to the current workbook. Quarterly facts prove
        # the same machinery can update an end-of-quarter run.
        for selected, status_prefix in ((latest_annual, "annual_compare"), (latest_quarter, "quarterly_ready")):
            if not selected:
                continue
            raw_value = selected["val"] * config.get("sign", 1)
            value_musd = round(raw_value / 1_000_000, 3)
            workbook_label = config.get("workbook_label", metric)
            current_value = find_current_value(ws, workbook_label) if status_prefix == "annual_compare" else None
            delta = None
            if isinstance(current_value, (int, float)):
                delta = round(value_musd - float(current_value), 3)
            period = selected["end"] if not selected.get("start") else f"{selected['start']}..{selected['end']}"
            rows.append(
                ExtractedFact(
                    metric=metric,
                    xbrl_tag=config["tag"],
                    value_usd=raw_value,
                    value_musd=value_musd,
                    period=period,
                    fiscal_year=int(selected["fy"]),
                    fiscal_period=str(selected["fp"]),
                    form=str(selected["form"]),
                    filed=str(selected["filed"]),
                    accession=str(selected["accn"]),
                    source_url=filing_url(str(selected["accn"])),
                    current_workbook_value_musd=current_value,
                    delta_musd=delta,
                    status=status_prefix if delta in (None, 0) else f"{status_prefix}_diff",
                )
            )
    return rows


def write_workbook_copy(source_path: Path, output_path: Path, rows: list[ExtractedFact]) -> None:
    shutil.copy2(source_path, output_path)
    wb = load_workbook(output_path)
    if "Automation_POC" in wb.sheetnames:
        del wb["Automation_POC"]
    ws = wb.create_sheet("Automation_POC", 0)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:L1"
    headers = [
        "metric",
        "period",
        "fy",
        "fp",
        "form",
        "value_musd",
        "current_workbook_value_musd",
        "delta_musd",
        "status",
        "xbrl_tag",
        "filed",
        "source_url",
    ]
    ws.append(headers)
    for fact in rows:
        ws.append(
            [
                fact.metric,
                fact.period,
                fact.fiscal_year,
                fact.fiscal_period,
                fact.form,
                fact.value_musd,
                fact.current_workbook_value_musd,
                fact.delta_musd,
                fact.status,
                fact.xbrl_tag,
                fact.filed,
                fact.source_url,
            ]
        )

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    widths = {
        "A": 28,
        "B": 24,
        "C": 8,
        "D": 8,
        "E": 10,
        "F": 14,
        "G": 24,
        "H": 12,
        "I": 24,
        "J": 48,
        "K": 12,
        "L": 80,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2):
        row[11].hyperlink = row[11].value
        row[11].style = "Hyperlink"
    wb.save(output_path)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--output-dir", type=Path, default=Path("outputs"))
    args = parser.parse_args()

    workbook_path = args.workbook.resolve()
    if not workbook_path.exists():
        print(f"Workbook not found: {workbook_path}", file=sys.stderr)
        return 2

    run_id = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    output_dir = (args.output_dir / f"goodyear_sec_dry_run_{run_id}").resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    company_facts = fetch_json(SEC_COMPANY_FACTS_URL)
    rows = build_rows(company_facts, workbook_path)
    payload = {
        "run_id": run_id,
        "entity": company_facts.get("entityName"),
        "source_api": SEC_COMPANY_FACTS_URL,
        "note": "Values are USD millions. Quarterly rows are not written into the main model; they prove extraction readiness.",
        "facts": [asdict(row) for row in rows],
    }

    json_path = output_dir / "goodyear_sec_dry_run.json"
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    workbook_copy = output_dir / f"{workbook_path.stem}_goodyear_sec_poc.xlsx"
    write_workbook_copy(workbook_path, workbook_copy, rows)

    diffs = [row for row in rows if row.status.endswith("_diff")]
    print(f"Wrote {json_path}")
    print(f"Wrote {workbook_copy}")
    print(f"Extracted facts: {len(rows)}; annual differences vs workbook: {len(diffs)}")
    for row in diffs:
        print(
            f"- {row.metric}: SEC {row.value_musd} vs workbook {row.current_workbook_value_musd} "
            f"(delta {row.delta_musd})"
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
