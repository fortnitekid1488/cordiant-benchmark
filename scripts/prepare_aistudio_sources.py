#!/usr/bin/env python3
"""Prepare latest-period source bundles for LLM extraction.

The output is meant for a non-technical workflow:
1. Use the dashboard's Prepare Sources action, or run this script directly.
2. Upload a batch folder's `FILES_FOR_AI_STUDIO` to the selected LLM provider.
3. StockAnalysis is the primary uploaded source; official files are fallback-only.
4. Use the dashboard's Build Excel action, or run `scripts/apply_aistudio_json.py`.
"""

from __future__ import annotations

import argparse
import html
import json
import math
import re
import shutil
import sys
import time
import urllib.error
import urllib.request
from datetime import date, datetime, timezone
from pathlib import Path
from urllib.parse import urljoin, urlparse, urlunparse

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
REGISTRY_PATH = ROOT / "config" / "company_source_registry.json"
PROVIDER_PROFILES_PATH = ROOT / "config" / "ai_provider_profiles.json"
SCHEMA_PATH = ROOT / "templates" / "aistudio_latest_quarter_schema.json"
PROMPT_TEMPLATE_PATH = ROOT / "templates" / "aistudio_prompt_template.txt"
DEFAULT_MAX_COMPANIES_PER_BATCH = 6
DEFAULT_MAX_ESTIMATED_INPUT_TOKENS = 64000
APPROX_CHARS_PER_TOKEN = 4
STOCKANALYSIS_STATEMENTS = (
    ("income_statement", "financials", "StockAnalysis income statement fallback."),
    ("balance_sheet", "financials/balance-sheet", "StockAnalysis balance sheet fallback."),
    ("cash_flow", "financials/cash-flow-statement", "StockAnalysis cash flow fallback."),
    ("employees", "employees", "StockAnalysis employee count table fallback."),
)
STOCKANALYSIS_FALLBACK_SOURCE_MODES = {
    ("michelin", "quarterly"),
}

METRICS = [
    "Total Revenues",
    "# employees",
    "Cost Of Revenues",
    "Gross Profit",
    "Gross Profit Margin %",
    "Other Operating Expenses, Total",
    "R&D Expenses",
    "Selling General & Admin Expenses",
    "Other Operating Expenses",
    "Operating Income",
    "EBIT Margin %",
    "EBITDA",
    "EBITDA Margin %",
    "Total Receivables",
    "Net Income",
    "Accounts Receivable, Total",
    "Inventory",
    "Total Current Liabilities",
    "Total Assets",
    "Total Equity",
    "Total Debt",
    "Capital Expenditure",
    "Levered Free Cash Flow",
    "Cash from Operations",
]

STOCKANALYSIS_ROW_MAP = {
    "Total Revenues": [("income_statement", "Revenue"), ("income_statement", "Operating Revenue"), ("income_statement", "Total Revenue")],
    "Cost Of Revenues": [("income_statement", "Cost of Revenue")],
    "Gross Profit": [("income_statement", "Gross Profit")],
    "Other Operating Expenses, Total": [("income_statement", "Operating Expenses")],
    "R&D Expenses": [("income_statement", "Research & Development")],
    "Selling General & Admin Expenses": [("income_statement", "Selling, General & Admin")],
    "Other Operating Expenses": [("income_statement", "Other Operating Expenses")],
    "Operating Income": [("income_statement", "Operating Income"), ("income_statement", "EBIT")],
    "EBITDA": [("income_statement", "EBITDA")],
    "Net Income": [("income_statement", "Net Income"), ("income_statement", "Net Income to Company")],
    "Total Receivables": [("balance_sheet", "Receivables")],
    "Accounts Receivable, Total": [("balance_sheet", "Accounts Receivable"), ("balance_sheet", "Receivables")],
    "Inventory": [("balance_sheet", "Inventory")],
    "Total Current Liabilities": [("balance_sheet", "Total Current Liabilities")],
    "Total Assets": [("balance_sheet", "Total Assets")],
    "Total Equity": [("balance_sheet", "Shareholders' Equity"), ("balance_sheet", "Total Common Equity"), ("balance_sheet", "Total Equity")],
    "Total Debt": [("balance_sheet", "Total Debt")],
    "Capital Expenditure": [("cash_flow", "Capital Expenditures")],
    "Levered Free Cash Flow": [("cash_flow", "Levered Free Cash Flow"), ("cash_flow", "Free Cash Flow")],
    "Cash from Operations": [("cash_flow", "Operating Cash Flow")],
}

COMPACT_SCHEMA_TEMPLATE = """{
  "reporting_scope": "{{REPORTING_SCOPE}}",
  "currency_rule": "Use company reporting currency; do not convert currencies.",
  "unit_rule": "Use millions. Convert billions to millions and note it.",
  "companies": [{
    "company_key": "string",
    "company_name": "string",
    "period_label": "string",
    "period_start": "YYYY-MM-DD or null",
    "period_end": "YYYY-MM-DD or null",
    "currency": "string",
    "unit": "millions",
    "facts": [{
      "metric": "requested metric name",
      "value": "number or null",
      "source_file": "uploaded file name or null",
      "source_url": "source URL or null",
      "source_type": "official_api | official_pdf | official_excel | official_ir_page | official_regulatory_portal | readable_aggregator_fallback | unknown",
      "page": "page/table page or null",
      "table_label": "table/section label or null",
      "evidence_quote": "short row/quote proving the value or null",
      "confidence": "high | medium | low",
      "review_required": true,
      "review_reason": "reason if review_required is true"
    }],
    "company_review_notes": ["string"]
  }]
}"""

MODE_CONFIG = {
    "quarterly": {
        "folder_prefix": "aistudio_quarterly_package",
        "mode_label": "quarterly_update",
        "reporting_scope": "latest reported quarter financial benchmark for tire manufacturers",
        "period_hint": "latest reported standalone quarter",
        "period_rule_1": "Use the latest reported quarter available in the sources for each company.",
        "period_rule_2": "If a source reports cumulative H1/9M/YTD numbers instead of a standalone quarter, say so in `period_label` and `company_review_notes`, and mark affected values `review_required`.",
        "period_instructions": (
            "Extract the newest reported quarterly period available in the uploaded files or listed URLs. "
            "Do not default to 2025 or Q1 2026 just because those examples appear in old files or URLs. "
            "Prefer standalone Q1/Q2/Q3/Q4 values. If only cumulative H1/9M/YTD is available, return it but mark review_required."
        ),
    },
    "annual": {
        "folder_prefix": "aistudio_annual_package",
        "mode_label": "annual_update",
        "reporting_scope": "latest full-year annual financial benchmark for tire manufacturers",
        "period_hint": "latest reported full fiscal year",
        "period_rule_1": "Use the latest full fiscal year or annual period available in the sources for each company.",
        "period_rule_2": "Do not use quarterly, interim, H1, 9M, YTD, TTM, or Q1/Q2/Q3/Q4 values for annual mode; if only those are available, set the metric to null and mark `review_required`.",
        "period_instructions": (
            "Extract the newest full-year annual period available in the uploaded files or listed URLs. "
            "Do not default to 2025 if a newer annual report exists in the sources. "
            "Ignore quarterly/interim/YTD-only documents for values in annual mode."
        ),
    },
}


def slugify(text: str) -> str:
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_") or "source"


def normalize_label(value: object) -> str:
    return " ".join(str(value or "").replace("\xa0", " ").strip().lower().split())


def source_filename(company_key: str, index: int, url: str, content_type: str | None) -> str:
    parsed = urlparse(url)
    path_name = Path(parsed.path).name
    suffix = Path(path_name).suffix.lower()
    if not suffix:
        if content_type and "pdf" in content_type:
            suffix = ".pdf"
        elif content_type and ("spreadsheet" in content_type or "excel" in content_type):
            suffix = ".xlsx"
        elif content_type and "json" in content_type:
            suffix = ".json"
        else:
            suffix = ".html"
    base = slugify(Path(path_name).stem if path_name else parsed.netloc)
    return f"{company_key}_{index:02d}_{base}{suffix}"


FAST_DOWNLOAD_TYPES = {"official_api", "official_pdf", "official_excel", "official_press_release"}


def load_provider_profiles() -> dict[str, dict]:
    return json.loads(PROVIDER_PROFILES_PATH.read_text(encoding="utf-8"))


def provider_label(provider_profile: dict | None) -> str:
    if not provider_profile:
        return "selected provider"
    return provider_profile.get("label") or "selected provider"


def provider_upload_limit(provider_profile: dict | None) -> int:
    if not provider_profile:
        return 0
    return int(provider_profile.get("max_upload_files_per_batch") or 0)


def schema_text_for_mode(mode: str) -> str:
    return COMPACT_SCHEMA_TEMPLATE.replace("{{REPORTING_SCOPE}}", MODE_CONFIG[mode]["reporting_scope"])


def period_hint_for_mode(mode: str) -> str:
    return MODE_CONFIG[mode]["period_hint"]


def is_quarter_specific_source(source: dict) -> bool:
    raw_text = f"{source.get('url', '')} {source.get('note', '')}".lower()
    normalized = re.sub(r"[^a-z0-9]+", " ", raw_text)
    patterns = (
        r"\bq[1-4]\b",
        r"\b[1-4]q\b",
        r"\b10\s*q\b",
        r"\bfirst\s+quarter\b",
        r"\bsecond\s+quarter\b",
        r"\bthird\s+quarter\b",
        r"\bfourth\s+quarter\b",
        r"\bthree\s+months\b",
        r"\b3\s+months\b",
        r"\bquarterly\b",
        r"\b31\s+march\b",
        r"\b30\s+june\b",
        r"\b30\s+september\b",
    )
    return any(re.search(pattern, normalized) for pattern in patterns)


def is_stockanalysis_source(source: dict) -> bool:
    return "stockanalysis.com" in str(source.get("url", "")).lower()


def is_deterministic_stockanalysis_source(source: dict) -> bool:
    return is_stockanalysis_source(source) and bool(source.get("stockanalysis_statement"))


def stockanalysis_statement_url(url: str, statement_path: str) -> str:
    parsed = urlparse(url)
    base_path = parsed.path.split("/financials", 1)[0].rstrip("/")
    path = f"{base_path}/{statement_path.strip('/')}/"
    query = "" if statement_path.strip("/") == "employees" else parsed.query
    return urlunparse(parsed._replace(path=path, query=query))


def expand_stockanalysis_source(source: dict) -> list[dict]:
    variants: list[dict] = []
    for statement_key, statement_path, note in STOCKANALYSIS_STATEMENTS:
        source_entry = dict(source)
        source_entry["url"] = stockanalysis_statement_url(str(source["url"]), statement_path)
        source_entry["stockanalysis_statement"] = statement_key
        source_entry["note"] = note
        variants.append(source_entry)
    return variants


def dedupe_sources(sources: list[dict]) -> list[dict]:
    seen: set[str] = set()
    result: list[dict] = []
    for source in sources:
        key = str(source.get("url") or "")
        if key in seen:
            continue
        seen.add(key)
        result.append(source)
    return result


def source_for_mode(source: dict, mode: str) -> dict | None:
    source_entry = dict(source)
    source_type = source_entry.get("type")
    if mode == "quarterly":
        if source_type in {"official_pdf", "official_press_release", "official_sec_filing"} and is_quarter_specific_source(source_entry):
            return None
        if source_type == "readable_aggregator_fallback":
            source_entry["note"] = "Readable fallback with latest quarterly financial statement tables."
        elif is_quarter_specific_source(source_entry):
            source_entry["note"] = "Official source page. Use the latest quarterly materials available there; do not treat older period examples as the target."
        return source_entry

    if source_type == "readable_aggregator_fallback":
        source_entry["url"] = source_entry["url"].replace("?p=quarterly", "")
        source_entry["note"] = "Readable fallback with annual financial statement tables."
        return source_entry

    if source_type in {"official_pdf", "official_press_release", "official_sec_filing"} and is_quarter_specific_source(source_entry):
        return None

    if is_quarter_specific_source(source_entry):
        source_entry["note"] = "Official source page. Use the latest annual/full-year materials available there; do not use quarterly examples as the target."

    return source_entry


def selected_sources_for_company(company: dict, mode: str) -> list[dict]:
    mode_sources = [
        mode_source
        for source in company.get("sources", [])
        if (mode_source := source_for_mode(source, mode)) is not None
    ]
    stockanalysis_sources = [
        expanded
        for source in mode_sources
        if is_stockanalysis_source(source)
        for expanded in expand_stockanalysis_source(source)
    ]
    stockanalysis_sources = dedupe_sources(stockanalysis_sources)
    if not stockanalysis_sources:
        return dedupe_sources(mode_sources)

    non_stockanalysis_sources = [
        source
        for source in mode_sources
        if not is_stockanalysis_source(source)
    ]
    if (company.get("key"), mode) in STOCKANALYSIS_FALLBACK_SOURCE_MODES:
        for source in non_stockanalysis_sources:
            source["note"] = (
                f"{source.get('note', '')} Use only when StockAnalysis does not provide the required target period."
            ).strip()
        return dedupe_sources([*stockanalysis_sources, *non_stockanalysis_sources])

    return stockanalysis_sources


def fetch_url(url: str, timeout: int = 6) -> tuple[bytes, str, str]:
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
            "Accept": "text/html,application/xhtml+xml,application/pdf,application/json,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
            "Accept-Language": "en-US,en;q=0.9",
        },
    )
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            return response.read(), response.headers.get("content-type", ""), response.url
    except urllib.error.HTTPError as exc:
        if exc.code == 404 and "/employees/" in url:
            fallback_url = url.replace("/employees/", "/company/")
            fallback_request = urllib.request.Request(fallback_url, headers=dict(request.header_items()))
            with urllib.request.urlopen(fallback_request, timeout=timeout) as response:
                return response.read(), response.headers.get("content-type", ""), response.url
        raise


def html_to_text(html: str) -> str:
    text = re.sub(r"<(script|style)[\s\S]*?</\1>", " ", html, flags=re.I)
    text = re.sub(r"<[^>]+>", " ", text)
    text = text.replace("&nbsp;", " ").replace("&amp;", "&")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def format_cell_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).replace("\n", " ").replace("\t", " ").strip()


def xlsx_to_text(workbook_path: Path, max_chars: int = 200000) -> str:
    """Convert official Excel downloads to a tab-separated text extract for provider upload.

    Several provider web UIs accept PDF/TXT/JSON more reliably than XLSX uploads.
    The original workbook is still kept in the package for provenance.
    """
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    sections: list[str] = [
        f"# Extracted text from {workbook_path.name}",
        "# Values are tab-separated. Blank trailing cells are removed from each row.",
        "",
    ]
    for ws in wb.worksheets:
        sections.append(f"## Sheet: {ws.title}")
        for row in ws.iter_rows(values_only=True):
            values = [format_cell_value(value) for value in row]
            while values and values[-1] == "":
                values.pop()
            if values:
                sections.append("\t".join(values))
        sections.append("")
    return "\n".join(sections)[:max_chars]


def pdf_to_text(pdf_path: Path, max_chars: int = 160000) -> str:
    """Convert uploaded official PDFs to a compact text extract for manual LLM review."""
    try:
        from pypdf import PdfReader
    except Exception as exc:  # noqa: BLE001 - dependency may be absent on a fresh machine.
        return f"# Could not extract text from {pdf_path.name}: {exc}"

    sections = [f"# Extracted text from {pdf_path.name}", ""]
    try:
        reader = PdfReader(pdf_path)
        for index, page in enumerate(reader.pages, start=1):
            text = page.extract_text() or ""
            text = re.sub(r"[ \t]+", " ", text).strip()
            if not text:
                continue
            sections.extend([f"## Page {index}", text, ""])
            if sum(len(section) for section in sections) >= max_chars:
                break
    except Exception as exc:  # noqa: BLE001 - keep package generation resilient.
        sections.append(f"Could not extract text: {exc}")

    return "\n".join(sections)[:max_chars]


def parse_stockanalysis_number(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value).strip()
    if not text or text in {"-", "—", "–"}:
        return None
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()")
    text = text.replace(",", "").replace("%", "").replace("−", "-")
    try:
        number = float(text)
    except ValueError:
        return None
    return -number if negative else number


def parse_stockanalysis_period_end(value: object) -> str | None:
    text = str(value or "")
    match = re.search(r"([A-Z][a-z]{2}\s+\d{1,2},\s+20\d{2})", text)
    if not match:
        return None
    try:
        return datetime.strptime(match.group(1), "%b %d, %Y").date().isoformat()
    except ValueError:
        return None


def parse_stockanalysis_date(value: object) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%b %d, %Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def stockanalysis_column_text(column: object) -> str:
    if isinstance(column, tuple):
        return " ".join(str(part or "") for part in column)
    return str(column or "")


def selected_stockanalysis_value_column(table: pd.DataFrame, mode: str) -> object:
    if mode != "annual":
        return table.columns[1]
    for column in table.columns[1:]:
        text = stockanalysis_column_text(column)
        if re.search(r"\bttm\b|trailing", text, flags=re.I):
            continue
        return column
    return table.columns[1]


def stockanalysis_table_from_source(path: Path, mode: str) -> tuple[pd.DataFrame, str, str | None]:
    table = pd.read_html(path)[0]
    first_col = table.columns[0]
    latest_col = selected_stockanalysis_value_column(table, mode)
    if isinstance(latest_col, tuple):
        period_label = str(latest_col[0])
        period_end = parse_stockanalysis_period_end(latest_col[1])
    else:
        period_label = str(latest_col)
        period_end = None
    if latest_col != table.columns[1]:
        table = table.loc[:, [first_col, latest_col]]
    return table, period_label, period_end


def stockanalysis_employees_from_source(
    path: Path,
    mode: str,
    target_period_end: str | None,
) -> tuple[float | None, str | None]:
    table = pd.read_html(path)[0]
    date_col = next((col for col in table.columns if normalize_label(col) == "date"), table.columns[0])
    employee_col = next((col for col in table.columns if normalize_label(col) == "employees"), None)
    if employee_col is None and len(table.columns) >= 2:
        label_col = table.columns[0]
        value_col = table.columns[1]
        for _, row in table.iterrows():
            if normalize_label(row[label_col]) == "employees":
                employee_count = parse_stockanalysis_number(row[value_col])
                if employee_count is not None:
                    return employee_count, "Employees from StockAnalysis company profile"
    if employee_col is None:
        return None, None

    target_date = parse_stockanalysis_date(target_period_end)
    first_valid: tuple[date | None, float, str] | None = None
    latest_annual: tuple[date | None, float, str] | None = None
    latest_before_target: tuple[date | None, float, str] | None = None

    for _, row in table.iterrows():
        employee_count = parse_stockanalysis_number(row[employee_col])
        if employee_count is None:
            continue
        row_date = parse_stockanalysis_date(row[date_col])
        label = f"Employees as of {row_date.isoformat() if row_date else row[date_col]}"
        if first_valid is None:
            first_valid = (row_date, employee_count, label)
        if row_date and row_date.month == 12 and latest_annual is None:
            latest_annual = (row_date, employee_count, label)
        if target_date and row_date and row_date <= target_date and latest_before_target is None:
            latest_before_target = (row_date, employee_count, label)

    selected = latest_before_target or (latest_annual if mode == "annual" else first_valid) or first_valid
    if not selected:
        return None, None
    _, employee_count, label = selected
    return employee_count, label


def period_label_has_quarterly_issue(label: str) -> bool:
    normalized = normalize_label(label)
    if re.search(r"\b(h1|1h|h2|2h|9m|6m|ytd|ttm)\b|half[-\s]?year|six\s+months|nine\s+months", normalized):
        return True
    return not bool(
        re.search(
            r"\b(q[1-4]|[1-4]q)\b|first\s+quarter|second\s+quarter|third\s+quarter|fourth\s+quarter|"
            r"\bquarterly\b|three\s+months|3\s+months",
            normalized,
        )
    )


def text_from_source(batch_dir: Path, source: dict) -> str:
    rel_path = source.get("text_extract_file") or source.get("downloaded_file")
    if not rel_path:
        return ""
    source_path = batch_dir / rel_path
    if not source_path.exists():
        return ""
    if source_path.suffix.lower() == ".pdf":
        try:
            from pypdf import PdfReader

            return "\n".join(page.extract_text() or "" for page in PdfReader(source_path).pages)
        except Exception:  # noqa: BLE001 - best-effort official-source extraction.
            return ""
    try:
        return source_path.read_text(encoding="utf-8", errors="ignore")
    except Exception:  # noqa: BLE001 - best-effort official-source extraction.
        return ""


def parse_michelin_quarterly_period(text: str) -> tuple[str, str] | None:
    if re.search(r"\b2026\s+Q1\b|\bQ1\s+2026\b|first\s+quarter", text, flags=re.I):
        return "Q1 2026", "2026-03-31"
    return None


def parse_michelin_quarterly_revenue(text: str) -> tuple[float, str] | None:
    exact_patterns = [
        r"Q1 revenue:.*?\b2025\s+Q1\b.*?\b2026\s+Q1\s+at\s+constant\s+FX\b.*?\b2026\s+Q1\s+at\s+current\s+FX\b",
        r"Group revenue:.*?\bQ1\s+2026\b.*?\bGroup\b",
    ]
    for pattern in exact_patterns:
        match = re.search(pattern, text, flags=re.I | re.S)
        if not match:
            continue
        numbers = [float(value.replace(",", "")) for value in re.findall(r"\b\d{1,3}(?:,\d{3})\b", match.group(0))]
        if numbers:
            return numbers[-1], "Q1 2026 sales evolution / Group revenue table: 2026 Q1 at current FX"

    rounded = re.search(r"Group revenue amounted to\s+€\s*([0-9]+(?:\.[0-9]+)?)\s+billion", text, flags=re.I)
    if rounded:
        return float(rounded.group(1)) * 1000, "Financial information for the three months ended March 31, 2026: Group revenue amounted to EUR billions"
    return None


def official_quarterly_fallback_company_json(
    batch_dir: Path,
    company: dict,
    manifest_company: dict,
    period_label: str,
    period_end: str | None,
    raw_values: dict[str, tuple[float | None, dict | None, str | None]],
    notes: list[str],
) -> dict | None:
    if company.get("key") != "michelin":
        return None

    official_texts: list[tuple[dict, str]] = []
    for source in manifest_company.get("sources", []):
        if source.get("type") == "readable_aggregator_fallback":
            continue
        text = text_from_source(batch_dir, source)
        if text:
            official_texts.append((source, text))
    if not official_texts:
        return None

    selected_period: tuple[str, str] | None = None
    revenue_fact: tuple[float, dict, str] | None = None
    for source, text in official_texts:
        selected_period = selected_period or parse_michelin_quarterly_period(text)
        parsed_revenue = parse_michelin_quarterly_revenue(text)
        if parsed_revenue:
            value, label = parsed_revenue
            revenue_fact = (value, source, label)
            if "sales evolution" in label:
                break
    if not selected_period or not revenue_fact:
        return None

    fallback_raw_values = {metric: (None, None, None) for metric in METRICS}
    value, source, row_label = revenue_fact
    fallback_raw_values["Total Revenues"] = (value, source, row_label)

    employee_value, employee_source, employee_label = raw_values.get("# employees", (None, None, None))
    if employee_value is not None:
        fallback_raw_values["# employees"] = (employee_value, employee_source, employee_label)

    facts = []
    for metric in METRICS:
        fact_value, fact_source, fact_label = fallback_raw_values.get(metric, (None, None, None))
        facts.append(fact_from_stockanalysis(metric, fact_value, fact_source, fact_label))

    official_notes = [
        *notes,
        (
            f"StockAnalysis selected `{period_label}` ({period_end or 'no period_end'}), which is not a standalone "
            "quarter. Used official Michelin Q1 material only for explicitly available quarterly facts."
        ),
        "Unavailable standalone quarterly metrics remain review_required instead of using H1/H2 values.",
    ]
    return {
        "company_key": company.get("key"),
        "company_name": company.get("display_name"),
        "period_label": selected_period[0],
        "period_start": None,
        "period_end": selected_period[1],
        "currency": company.get("currency"),
        "unit": "millions",
        "facts": facts,
        "company_review_notes": official_notes,
    }


def stockanalysis_lookup(table: pd.DataFrame, labels: list[str]) -> tuple[float | None, str | None]:
    first_col = table.columns[0]
    latest_col = table.columns[1]
    label_map = {str(row[first_col]).strip(): row[latest_col] for _, row in table.iterrows()}
    normalized_map = {normalize_label(label): (label, value) for label, value in label_map.items()}
    for label in labels:
        exact = label_map.get(label)
        if exact is not None:
            return parse_stockanalysis_number(exact), label
        normalized = normalized_map.get(normalize_label(label))
        if normalized:
            actual_label, actual_value = normalized
            return parse_stockanalysis_number(actual_value), actual_label
    return None, None


def fact_from_stockanalysis(
    metric: str,
    value: float | None,
    source: dict | None,
    row_label: str | None,
    computed_label: str | None = None,
) -> dict:
    has_value = value is not None
    source_file = None
    if source:
        source_file = Path(source.get("ai_studio_upload_file") or source.get("downloaded_file") or "").name or None
    evidence = computed_label or (f"{row_label}: {value}" if row_label and has_value else None)
    return {
        "metric": metric,
        "value": value,
        "source_file": source_file,
        "source_url": (source or {}).get("final_url") or (source or {}).get("url"),
        "source_type": (source or {}).get("type") or "readable_aggregator_fallback",
        "page": None,
        "table_label": computed_label or row_label,
        "evidence_quote": evidence,
        "confidence": "high" if has_value else "low",
        "review_required": not has_value,
        "review_reason": None if has_value else "Metric not found in the latest StockAnalysis statement table.",
    }


def build_stockanalysis_company_json(batch_dir: Path, company: dict, mode: str) -> dict | None:
    statement_tables: dict[str, pd.DataFrame] = {}
    statement_sources: dict[str, dict] = {}
    period_labels: list[str] = []
    period_ends: list[str] = []
    employee_sources: list[tuple[dict, Path]] = []
    notes: list[str] = []

    for source in company.get("sources", []):
        statement = source.get("stockanalysis_statement")
        if not statement:
            continue
        downloaded_file = source.get("downloaded_file")
        if not downloaded_file:
            notes.append(f"{statement}: source was not downloaded.")
            continue
        source_path = batch_dir / downloaded_file
        if statement == "employees":
            employee_sources.append((source, source_path))
            continue
        try:
            table, period_label, period_end = stockanalysis_table_from_source(source_path, mode)
        except Exception as exc:  # noqa: BLE001 - record in generated JSON notes.
            notes.append(f"{statement}: failed to parse StockAnalysis table: {exc}")
            continue
        statement_tables[statement] = table
        statement_sources[statement] = source
        if period_label:
            period_labels.append(period_label)
        if period_end:
            period_ends.append(period_end)

    if not statement_tables:
        return None

    period_label = period_labels[0] if period_labels else ""
    period_end = period_ends[0] if period_ends else None
    mismatched_labels = sorted(set(period_labels))
    mismatched_ends = sorted(set(period_ends))
    if len(mismatched_labels) > 1 or len(mismatched_ends) > 1:
        notes.append(
            "StockAnalysis statements disagree on latest period: "
            f"labels={mismatched_labels}, period_ends={mismatched_ends}."
        )

    raw_values: dict[str, tuple[float | None, dict | None, str | None]] = {}
    for metric, candidates in STOCKANALYSIS_ROW_MAP.items():
        raw_values[metric] = (None, None, None)
        for statement, row_label in candidates:
            table = statement_tables.get(statement)
            if table is None:
                continue
            value, actual_label = stockanalysis_lookup(table, [row_label])
            if value is not None:
                raw_values[metric] = (value, statement_sources.get(statement), actual_label)
                break

    if employee_sources:
        raw_values["# employees"] = (None, None, None)
        for source, source_path in employee_sources:
            try:
                value, employee_label = stockanalysis_employees_from_source(source_path, mode, period_end)
            except Exception as exc:  # noqa: BLE001 - record in generated JSON notes.
                notes.append(f"employees: failed to parse StockAnalysis employee table: {exc}")
                continue
            if value is not None:
                raw_values["# employees"] = (value, source, employee_label)
                break

    if mode == "quarterly" and period_label_has_quarterly_issue(period_label):
        official_fallback = official_quarterly_fallback_company_json(
            batch_dir,
            company,
            company,
            period_label,
            period_end,
            raw_values,
            notes,
        )
        if official_fallback:
            return official_fallback

    revenue = raw_values.get("Total Revenues", (None, None, None))[0]
    gross_profit = raw_values.get("Gross Profit", (None, None, None))[0]
    operating_income = raw_values.get("Operating Income", (None, None, None))[0]
    ebitda = raw_values.get("EBITDA", (None, None, None))[0]
    if revenue:
        income_source = statement_sources.get("income_statement")
        raw_values["Gross Profit Margin %"] = (
            round(gross_profit / revenue, 6) if gross_profit is not None else None,
            income_source,
            "computed: Gross Profit / Revenue",
        )
        raw_values["EBIT Margin %"] = (
            round(operating_income / revenue, 6) if operating_income is not None else None,
            income_source,
            "computed: Operating Income / Revenue",
        )
        raw_values["EBITDA Margin %"] = (
            round(ebitda / revenue, 6) if ebitda is not None else None,
            income_source,
            "computed: EBITDA / Revenue",
        )

    facts = []
    for metric in METRICS:
        value, source, row_label = raw_values.get(metric, (None, None, None))
        computed_label = row_label if row_label and row_label.startswith("computed:") else None
        facts.append(fact_from_stockanalysis(metric, value, source, row_label, computed_label))

    return {
        "company_key": company.get("key"),
        "company_name": company.get("display_name"),
        "period_label": period_label,
        "period_start": None,
        "period_end": period_end,
        "currency": company.get("currency"),
        "unit": "millions",
        "facts": facts,
        "company_review_notes": notes
        + [
            "Diagnostic draft from freshly downloaded StockAnalysis statement tables.",
            "Do not treat this draft as the final dashboard JSON unless it is reviewed or replaced by an LLM extraction.",
        ],
    }


def write_stockanalysis_auto_json(output_dir: Path, batch_dir: Path, batch_manifest: dict, mode: str) -> None:
    companies = []
    for company in batch_manifest.get("companies", []):
        company_json = build_stockanalysis_company_json(batch_dir, company, mode)
        if company_json:
            companies.append(company_json)
    if not companies:
        return
    json_dir = output_dir / "stockanalysis_auto_json"
    json_dir.mkdir(exist_ok=True)
    payload = {
        "companies": companies,
        "auto_generated": {
            "source": "stockanalysis_direct_parser_draft",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "mode": mode,
            "batch_id": batch_manifest.get("batch_id"),
        },
    }
    (json_dir / f"{batch_manifest['batch_id']}.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


SEC_RELEVANT_TAGS = {
    "RevenueFromContractWithCustomerExcludingAssessedTax",
    "CostOfRevenue",
    "CostOfGoodsAndServicesSold",
    "GrossProfit",
    "SellingGeneralAndAdministrativeExpense",
    "SellingAndMarketingExpense",
    "GeneralAndAdministrativeExpense",
    "OperatingIncomeLoss",
    "NetIncomeLoss",
    "AccountsReceivableNetCurrent",
    "AccountsReceivableNet",
    "InventoryNet",
    "InventoryFinishedGoodsNetOfReserves",
    "LiabilitiesCurrent",
    "Assets",
    "AssetsCurrent",
    "StockholdersEquity",
    "StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest",
    "LongTermDebtCurrent",
    "LongTermDebtNoncurrent",
    "ShortTermBorrowings",
    "LongTermDebtAndFinanceLeaseObligationsCurrent",
    "LongTermDebtAndFinanceLeaseObligationsNoncurrent",
    "OperatingLeaseLiability",
    "FinanceLeaseLiability",
    "PaymentsToAcquirePropertyPlantAndEquipment",
    "NetCashProvidedByUsedInOperatingActivities",
    "DepreciationDepletionAndAmortization",
    "DepreciationDepletionAndAmortizationExpense",
}


def slim_sec_company_facts(payload: bytes, max_facts_per_unit: int = 16) -> dict:
    raw = json.loads(payload)
    us_gaap = raw.get("facts", {}).get("us-gaap", {})
    tags = {}
    for tag in sorted(SEC_RELEVANT_TAGS):
        tag_payload = us_gaap.get(tag)
        if not tag_payload:
            continue
        units = {}
        for unit_name, facts in tag_payload.get("units", {}).items():
            relevant = [
                {
                    "fy": fact.get("fy"),
                    "fp": fact.get("fp"),
                    "form": fact.get("form"),
                    "start": fact.get("start"),
                    "end": fact.get("end"),
                    "filed": fact.get("filed"),
                    "accn": fact.get("accn"),
                    "val": fact.get("val"),
                }
                for fact in facts
                if fact.get("form") in {"10-K", "10-Q"}
            ]
            relevant.sort(
                key=lambda fact: (
                    str(fact.get("end") or ""),
                    str(fact.get("filed") or ""),
                    str(fact.get("start") or ""),
                ),
                reverse=True,
            )
            if relevant:
                units[unit_name] = relevant[:max_facts_per_unit]
        if units:
            tags[tag] = {
                "label": tag_payload.get("label"),
                "description": tag_payload.get("description"),
                "units": units,
            }
    return {
        "source": "SEC Company Facts compact extract",
        "entityName": raw.get("entityName"),
        "cik": raw.get("cik"),
        "note": "Compact extract for provider upload. Monetary values are in raw SEC units; convert to millions in the answer.",
        "tags": tags,
    }


def discover_michelin_quarterly_pdf_sources(source: dict, payload: bytes, final_url: str, mode: str) -> list[dict]:
    if mode != "quarterly" or "michelin.com" not in final_url.lower():
        return []
    html_text = payload.decode("utf-8", "ignore")
    discovered: list[dict] = []
    seen: set[str] = set()
    for tag_match in re.finditer(r"<a\b[^>]*href=[\"']([^\"']+)[\"'][^>]*>", html_text, flags=re.I | re.S):
        href = html.unescape(tag_match.group(1))
        tag_text = html.unescape(tag_match.group(0))
        haystack = f"{href} {tag_text}".lower()
        if ".pdf" not in haystack or "invitation" in haystack:
            continue
        is_q1_material = any(
            marker in haystack
            for marker in (
                "1st-quarter",
                "first-quarter",
                "first quarter",
                "march-31-2026",
                "march 31 2026",
            )
        )
        if not is_q1_material:
            continue
        url = urljoin(final_url, href)
        if url in seen:
            continue
        seen.add(url)
        discovered.append(
            {
                "type": "official_pdf",
                "url": url,
                "note": "Official Michelin Q1 PDF discovered from the results and sales page for standalone quarterly review.",
            }
        )
        if len(discovered) >= 1:
            break
    return discovered


def materialize_source(
    source: dict,
    batch_dir: Path,
    sources_dir: Path,
    company_key: str,
    source_index: int,
) -> tuple[dict, bytes | None]:
    source_entry = dict(source)
    try:
        payload, content_type, final_url = fetch_url(source["url"])
        filename = source_filename(company_key, source_index, final_url, content_type)
        file_path = sources_dir / filename
        file_path.write_bytes(payload)
        source_entry.update(
            {
                "download_status": "downloaded",
                "downloaded_file": str(file_path.relative_to(batch_dir)),
                "content_type": content_type,
                "final_url": final_url,
                "bytes": len(payload),
            }
        )
        if "html" in content_type.lower() or filename.endswith(".html"):
            text_path = file_path.with_suffix(".txt")
            text_path.write_text(html_to_text(payload.decode("utf-8", "ignore"))[:200000], encoding="utf-8")
            source_entry["text_extract_file"] = str(text_path.relative_to(batch_dir))
            source_entry["ai_studio_upload_file"] = str(text_path.relative_to(batch_dir))
        if "pdf" in content_type.lower() or filename.endswith(".pdf"):
            text_path = file_path.with_suffix(".txt")
            text_path.write_text(pdf_to_text(file_path), encoding="utf-8")
            source_entry["text_extract_file"] = str(text_path.relative_to(batch_dir))
            source_entry["ai_studio_upload_file"] = str(text_path.relative_to(batch_dir))
            source_entry["ai_studio_upload_note"] = (
                "Upload this compact text extract instead of the raw PDF for manual LLM review."
            )
        if filename.endswith(".xlsx"):
            text_path = file_path.with_suffix(".txt")
            text_path.write_text(xlsx_to_text(file_path), encoding="utf-8")
            source_entry["text_extract_file"] = str(text_path.relative_to(batch_dir))
            source_entry["ai_studio_upload_file"] = str(text_path.relative_to(batch_dir))
            source_entry["ai_studio_upload_note"] = (
                "Some provider web UIs may reject XLSX files. Upload this text_extract_file instead."
            )
        if source.get("type") == "official_api" and "data.sec.gov/api/xbrl/companyfacts" in final_url:
            slim_path = file_path.with_name(f"{file_path.stem}_slim.json")
            slim_path.write_text(
                json.dumps(slim_sec_company_facts(payload), ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            source_entry["slim_extract_file"] = str(slim_path.relative_to(batch_dir))
            source_entry["ai_studio_upload_file"] = str(slim_path.relative_to(batch_dir))
            source_entry["ai_studio_upload_note"] = (
                "Upload this compact SEC extract instead of the full Company Facts JSON."
            )
        time.sleep(0.5)
        return source_entry, payload
    except Exception as exc:  # noqa: BLE001 - record for user/AI Studio
        source_entry.update({"download_status": "failed", "error": str(exc)})
        return source_entry, None


def llm_upload_file_for_source(source: dict) -> str | None:
    if source.get("llm_upload_skip"):
        return None
    upload_file = source.get("ai_studio_upload_file")
    if not upload_file and source.get("downloaded_file", "").endswith(".xlsx"):
        upload_file = source.get("text_extract_file")
    if not upload_file:
        upload_file = source.get("downloaded_file")
    return upload_file or None


def upload_files_for_batch(batch_manifest: dict) -> list[str]:
    files: list[str] = []
    for company in batch_manifest["companies"]:
        for source in company["sources"]:
            upload_file = llm_upload_file_for_source(source)
            if upload_file:
                files.append(upload_file)
    return files


def upload_files_for_company(company_manifest: dict) -> list[str]:
    files: list[str] = []
    for source in company_manifest.get("sources", []):
        upload_file = llm_upload_file_for_source(source)
        if upload_file:
            files.append(upload_file)
    return files


def estimated_upload_files_for_company(company: dict, mode: str, download_mode: str) -> int:
    count = 0
    for mode_source in selected_sources_for_company(company, mode):
        if download_mode == "fast" and mode_source.get("type") not in FAST_DOWNLOAD_TYPES:
            continue
        count += 1
    return count


def compact_source_manifest(batch_manifest: dict) -> dict:
    compact = {
        "batch_id": batch_manifest.get("batch_id"),
        "title": batch_manifest.get("title"),
        "companies": [],
    }
    for company in batch_manifest["companies"]:
        compact_company = {
            "key": company.get("key"),
            "display_name": company.get("display_name"),
            "country": company.get("country"),
            "currency": company.get("currency"),
            "period_target": company.get("period_target"),
            "sources": [],
        }
        for source in company.get("sources", []):
            upload_file = llm_upload_file_for_source(source)
            note = source.get("note") or source.get("ai_studio_upload_note") or source.get("note_for_user")
            if is_deterministic_stockanalysis_source(source):
                note_prefix = f"{note} " if note else ""
                note = (
                    f"{note_prefix}Primary LLM upload source for standard benchmark metrics."
                ).strip()
            compact_company["sources"].append(
                {
                    "source_type": source.get("type"),
                    "upload_file": Path(upload_file).name if upload_file else None,
                    "source_url": source.get("final_url") or source.get("url"),
                    "download_status": source.get("download_status"),
                    "note": note,
                }
            )
        compact["companies"].append(compact_company)
    return compact


def write_upload_list(batch_dir: Path, batch_manifest: dict) -> None:
    files = upload_files_for_batch(batch_manifest)
    lines = [
        "# Upload exactly these source files to Qwen Studio or the selected provider for this batch.",
        "# StockAnalysis files are the primary sources. Official files appear only as fallback when StockAnalysis is missing the target period or company.",
        "# Do not upload aistudio_latest_quarter_schema.json or source_manifest.json; the prompt already includes them.",
        "# Do not upload raw .xlsx/PDF files or full SEC Company Facts JSON when an extract is listed.",
        "",
        *files,
        "",
    ]
    (batch_dir / "FILES_TO_UPLOAD.txt").write_text("\n".join(lines), encoding="utf-8")
    write_upload_folder(batch_dir, files)


def write_upload_folder(batch_dir: Path, files: list[str]) -> None:
    upload_dir = batch_dir / "FILES_FOR_AI_STUDIO"
    if upload_dir.exists():
        shutil.rmtree(upload_dir)
    upload_dir.mkdir()
    used_names: set[str] = set()
    for rel_path in files:
        source_path = (batch_dir / rel_path).resolve()
        if not source_path.is_file() or batch_dir.resolve() not in source_path.parents:
            continue
        target_name = source_path.name
        if target_name in used_names:
            target_name = f"{slugify(source_path.parent.name)}_{target_name}"
        used_names.add(target_name)
        shutil.copy2(source_path, upload_dir / target_name)


def estimated_source_chars(source_path: Path) -> int:
    if source_path.suffix.lower() == ".pdf":
        try:
            from pypdf import PdfReader

            reader = PdfReader(source_path)
            text = "\n".join(page.extract_text() or "" for page in reader.pages)
            if text.strip():
                return len(text)
        except Exception:  # noqa: BLE001 - estimate only; package generation should continue.
            pass
        return max(1, source_path.stat().st_size // 8)
    return source_path.stat().st_size


def batch_input_estimate(batch_dir: Path) -> dict[str, int]:
    prompt_path = batch_dir / "prompt_for_aistudio.txt"
    prompt_chars = prompt_path.stat().st_size if prompt_path.exists() else 0
    upload_bytes = 0
    source_chars = 0
    upload_count = 0
    for rel_path in upload_files_from_written_list(batch_dir):
        source_path = batch_dir / rel_path
        if source_path.is_file():
            upload_bytes += source_path.stat().st_size
            source_chars += estimated_source_chars(source_path)
            upload_count += 1
    total_chars = prompt_chars + source_chars
    return {
        "prompt_chars": prompt_chars,
        "upload_bytes": upload_bytes,
        "estimated_source_chars": source_chars,
        "upload_count": upload_count,
        "approx_input_tokens": math.ceil(total_chars / APPROX_CHARS_PER_TOKEN),
    }


def upload_files_from_written_list(batch_dir: Path) -> list[Path]:
    upload_list = batch_dir / "FILES_TO_UPLOAD.txt"
    if not upload_list.exists():
        return []
    files = []
    for line in upload_list.read_text(encoding="utf-8").splitlines():
        rel_path = line.strip()
        if rel_path and not rel_path.startswith("#"):
            files.append(Path(rel_path))
    return files


def write_manifest(batch_dir: Path, batch_manifest: dict) -> None:
    (batch_dir / "source_manifest.json").write_text(
        json.dumps(batch_manifest, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def prepare_batch(
    batch: dict,
    company_lookup: dict[str, dict],
    output_dir: Path,
    schema_text: str,
    prompt_template: str,
    mode: str,
    download_mode: str,
    provider: str,
    provider_profile: dict,
) -> dict:
    batch_dir = output_dir / batch["batch_id"]
    sources_dir = batch_dir / "sources"
    sources_dir.mkdir(parents=True, exist_ok=True)
    batch_manifest: dict = {
        "batch_id": batch["batch_id"],
        "title": batch["title"],
        "provider": {
            "id": provider,
            "label": provider_label(provider_profile),
            "max_upload_files_per_batch": provider_upload_limit(provider_profile),
            "limit_summary": provider_profile.get("limit_summary"),
        },
        "companies": [],
    }

    for company_key in batch["companies"]:
        company = company_lookup[company_key]
        company_entry = {
            "key": company["key"],
            "display_name": company["display_name"],
            "country": company["country"],
            "currency": company["currency"],
            "period_target": period_hint_for_mode(mode),
            "sources": [],
        }
        mode_sources = selected_sources_for_company(company, mode)
        for i, source in enumerate(mode_sources, start=1):
            source_entry = dict(source)
            if download_mode == "fast" and source.get("type") not in FAST_DOWNLOAD_TYPES:
                source_entry.update(
                    {
                        "download_status": "skipped_fast_mode",
                        "note_for_user": "URL is included in the manifest/prompt, but the fast preparer did not download this web page.",
                    }
                )
                company_entry["sources"].append(source_entry)
                continue
            source_entry, payload = materialize_source(source, batch_dir, sources_dir, company["key"], i)
            company_entry["sources"].append(source_entry)
            if payload and source_entry.get("content_type") and "html" in str(source_entry["content_type"]).lower():
                discovered_sources = discover_michelin_quarterly_pdf_sources(
                    source_entry,
                    payload,
                    str(source_entry.get("final_url") or source_entry.get("url")),
                    mode,
                )
                if discovered_sources and source_entry.get("type") == "official_ir_page":
                    source_entry["llm_upload_skip"] = True
                    source_entry["ai_studio_upload_file"] = None
                    source_entry["ai_studio_upload_note"] = (
                        "Discovery page used to find the latest official PDF; do not upload it when the PDF is listed."
                    )
                for discovered_source in discovered_sources:
                    discovered_index = len(company_entry["sources"]) + 1
                    discovered_entry, _ = materialize_source(
                        discovered_source,
                        batch_dir,
                        sources_dir,
                        company["key"],
                        discovered_index,
                    )
                    company_entry["sources"].append(discovered_entry)
        batch_manifest["companies"].append(company_entry)

    write_manifest(batch_dir, batch_manifest)

    company_list = "\n".join(
        f"- {company_lookup[key]['display_name']} ({company_lookup[key]['currency']}, {period_hint_for_mode(mode)})"
        for key in batch["companies"]
    )
    metric_list = "\n".join(f"- {metric}" for metric in METRICS)
    prompt = (
        prompt_template.replace("{{COMPANY_LIST}}", company_list)
        .replace("{{METRIC_LIST}}", metric_list)
        .replace("{{PERIOD_MODE}}", MODE_CONFIG[mode]["mode_label"])
        .replace("{{PERIOD_RULE_1}}", MODE_CONFIG[mode]["period_rule_1"])
        .replace("{{PERIOD_RULE_2}}", MODE_CONFIG[mode]["period_rule_2"])
        .replace("{{PERIOD_INSTRUCTIONS}}", MODE_CONFIG[mode]["period_instructions"])
        .replace("{{SCHEMA_JSON}}", schema_text_for_mode(mode))
        .replace("{{SOURCE_MANIFEST}}", json.dumps(compact_source_manifest(batch_manifest), ensure_ascii=False, indent=2))
    )
    (batch_dir / "prompt_for_aistudio.txt").write_text(prompt, encoding="utf-8")
    shutil.copy2(SCHEMA_PATH, batch_dir / "aistudio_latest_quarter_schema.json")
    write_upload_list(batch_dir, batch_manifest)
    batch_manifest["input_estimate"] = batch_input_estimate(batch_dir)
    max_upload_files = provider_upload_limit(provider_profile)
    upload_count = batch_manifest["input_estimate"].get("upload_count", 0)
    if max_upload_files > 0 and upload_count > max_upload_files:
        oversize_companies = [
            {
                "key": company.get("key"),
                "display_name": company.get("display_name"),
                "upload_count": len(upload_files_for_company(company)),
            }
            for company in batch_manifest["companies"]
            if len(upload_files_for_company(company)) > max_upload_files
        ]
        batch_manifest["provider_limit_warning"] = {
            "message": (
                f"Batch has {upload_count} upload files, above {provider_label(provider_profile)} "
                f"limit of {max_upload_files}. Split further or reduce sources."
            ),
            "oversize_companies": oversize_companies,
        }
    write_manifest(batch_dir, batch_manifest)
    write_stockanalysis_auto_json(output_dir, batch_dir, batch_manifest, mode)
    return batch_manifest


def split_one_batch(batch: dict, max_companies_per_batch: int) -> list[dict]:
    if max_companies_per_batch <= 0 or len(batch["companies"]) <= max_companies_per_batch:
        return [batch]
    split: list[dict] = []
    companies = batch["companies"]
    for index in range(0, len(companies), max_companies_per_batch):
        part = index // max_companies_per_batch + 1
        split.append(
            {
                "batch_id": f"{batch['batch_id']}_part_{part}",
                "title": f"{batch['title']} part {part}",
                "companies": companies[index : index + max_companies_per_batch],
            }
        )
    return split


def split_batches(batches: list[dict], max_companies_per_batch: int) -> list[dict]:
    split: list[dict] = []
    for batch in batches:
        split.extend(split_one_batch(batch, max_companies_per_batch))
    return split


def split_one_batch_by_upload_limit(
    batch: dict,
    company_lookup: dict[str, dict],
    mode: str,
    download_mode: str,
    max_upload_files: int,
) -> list[dict]:
    if max_upload_files <= 0:
        return [batch]

    parts: list[list[str]] = []
    current: list[str] = []
    current_count = 0
    for company_key in batch["companies"]:
        company_count = estimated_upload_files_for_company(company_lookup[company_key], mode, download_mode)
        if current and current_count + company_count > max_upload_files:
            parts.append(current)
            current = []
            current_count = 0
        current.append(company_key)
        current_count += company_count
    if current:
        parts.append(current)

    if len(parts) <= 1:
        return [batch]

    return [
        {
            "batch_id": f"{batch['batch_id']}_part_{index}",
            "title": f"{batch['title']} part {index}",
            "companies": companies,
        }
        for index, companies in enumerate(parts, start=1)
    ]


def split_batches_by_upload_limit(
    batches: list[dict],
    company_lookup: dict[str, dict],
    mode: str,
    download_mode: str,
    max_upload_files: int,
) -> list[dict]:
    split: list[dict] = []
    for batch in batches:
        split.extend(split_one_batch_by_upload_limit(batch, company_lookup, mode, download_mode, max_upload_files))
    return split


def split_batch_in_half(batch: dict) -> list[dict]:
    companies = batch["companies"]
    split_size = math.ceil(len(companies) / 2)
    return split_one_batch(batch, split_size)


def prepare_batch_with_budget(
    batch: dict,
    company_lookup: dict[str, dict],
    output_dir: Path,
    schema_text: str,
    prompt_template: str,
    mode: str,
    download_mode: str,
    max_estimated_input_tokens: int,
    max_upload_files: int,
    provider: str,
    provider_profile: dict,
) -> list[dict]:
    batch_manifest = prepare_batch(
        batch,
        company_lookup,
        output_dir,
        schema_text,
        prompt_template,
        mode,
        download_mode,
        provider,
        provider_profile,
    )
    estimate = batch_manifest.get("input_estimate", {})
    token_limit_exceeded = (
        max_estimated_input_tokens > 0
        and estimate.get("approx_input_tokens", 0) > max_estimated_input_tokens
    )
    file_limit_exceeded = (
        max_upload_files > 0
        and estimate.get("upload_count", 0) > max_upload_files
    )
    if (token_limit_exceeded or file_limit_exceeded) and len(batch["companies"]) > 1:
        shutil.rmtree(output_dir / batch["batch_id"])
        manifests: list[dict] = []
        if file_limit_exceeded:
            smaller_batches = split_one_batch_by_upload_limit(
                batch,
                company_lookup,
                mode,
                download_mode,
                max_upload_files,
            )
            if len(smaller_batches) == 1:
                smaller_batches = split_batch_in_half(batch)
        else:
            smaller_batches = split_batch_in_half(batch)
        for smaller_batch in smaller_batches:
            manifests.extend(
                prepare_batch_with_budget(
                    smaller_batch,
                    company_lookup,
                    output_dir,
                    schema_text,
                    prompt_template,
                    mode,
                    download_mode,
                    max_estimated_input_tokens,
                    max_upload_files,
                    provider,
                    provider_profile,
                )
            )
        return manifests
    return [batch_manifest]


def write_package_metadata(output_dir: Path, batch_manifests: list[dict], mode: str, provider: str, provider_profile: dict) -> None:
    metadata = {
        "mode": mode,
        "provider": {
            "id": provider,
            "label": provider_label(provider_profile),
            "site_label": provider_profile.get("site_label"),
            "max_upload_files_per_batch": provider_upload_limit(provider_profile),
            "limit_summary": provider_profile.get("limit_summary"),
            "limit_confidence": provider_profile.get("limit_confidence"),
        },
        "batch_count": len(batch_manifests),
        "batches": [
            {
                "batch_id": manifest.get("batch_id"),
                "upload_count": manifest.get("input_estimate", {}).get("upload_count", 0),
                "approx_input_tokens": manifest.get("input_estimate", {}).get("approx_input_tokens", 0),
            }
            for manifest in batch_manifests
        ],
    }
    (output_dir / "package_metadata.json").write_text(json.dumps(metadata, ensure_ascii=False, indent=2), encoding="utf-8")


def write_instructions(output_dir: Path, batch_manifests: list[dict], mode: str, provider_profile: dict) -> None:
    max_upload_files = provider_upload_limit(provider_profile)
    upload_limit_text = (
        f"Selected provider: {provider_label(provider_profile)}. Each batch is split to at most {max_upload_files} upload files."
        if max_upload_files > 0
        else f"Selected provider: {provider_label(provider_profile)}. No hard file-count split is applied; batches are controlled by input-token estimate."
    )
    lines = [
        f"# LLM {MODE_CONFIG[mode]['mode_label']} Package",
        "",
        upload_limit_text,
        "",
        "StockAnalysis files are the primary uploaded source for standard income statement, balance sheet, cash-flow, and employee rows.",
        "Official/company files are included only when StockAnalysis is unavailable or does not expose the required target period.",
        "A diagnostic local parser draft may exist in `stockanalysis_auto_json/`, but final dashboard input belongs in `aistudio_json/` after LLM review.",
        "",
        "1. Open one batch folder.",
        "2. In the provider web UI, turn off web search / grounding for this extraction run if that option is enabled.",
        "3. Open `FILES_FOR_AI_STUDIO` and upload every file in that folder.",
        "   That folder contains only the source attachments for the selected model.",
        "   It already replaces raw `.xlsx`, `.pdf`, and oversized SEC JSON files with upload-friendly extracts.",
        "   Do not upload `source_manifest.json`, the schema JSON, or `FILES_TO_UPLOAD.txt`; the prompt already includes the needed structure.",
        "   If some sources are marked `skipped_fast_mode` or `failed`, the prompt still contains their URLs. The model can use them as references, or you can manually download/upload those files later.",
        "4. Copy-paste `prompt_for_aistudio.txt` into the provider web UI.",
        "5. The model must return raw JSON only.",
        "6. Save each answer as `aistudio_json/<batch_folder_name>.json`, for example:",
    ]
    lines.extend(f"   - `aistudio_json/{manifest['batch_id']}.json`" for manifest in batch_manifests)
    lines.extend([
        "7. Run the matching update command from the project folder.",
        "",
        "Do not edit the original workbook. The update script creates a new workbook copy.",
        "",
        "## Batch Download Summary",
        "",
    ])
    for manifest in batch_manifests:
        downloaded = sum(1 for c in manifest["companies"] for s in c["sources"] if s.get("download_status") == "downloaded")
        failed = sum(1 for c in manifest["companies"] for s in c["sources"] if s.get("download_status") == "failed")
        estimate = manifest.get("input_estimate", {})
        approx_tokens = estimate.get("approx_input_tokens")
        estimate_text = f", approx input {approx_tokens:,} tokens" if approx_tokens else ""
        upload_count = estimate.get("upload_count", 0)
        limit_warning = " WARNING: above provider file limit" if manifest.get("provider_limit_warning") else ""
        lines.append(f"- `{manifest['batch_id']}`: {upload_count} upload files, downloaded {downloaded}, failed {failed}{estimate_text}{limit_warning}")
    lines.append("")
    (output_dir / "README_FOR_USER.md").write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output-dir", type=Path, default=ROOT / "outputs")
    parser.add_argument("--mode", choices=sorted(MODE_CONFIG), default="quarterly")
    parser.add_argument("--workbook", type=Path, default=ROOT / "Бенч финансовой отчетности_мэйджоры.xlsx")
    parser.add_argument("--download-mode", choices=["fast", "full"], default="fast")
    provider_profiles = load_provider_profiles()
    parser.add_argument("--provider", choices=sorted(provider_profiles), default="qwen")
    parser.add_argument(
        "--max-upload-files-per-batch",
        type=int,
        help="Override selected provider file-count limit. Use 0 to disable file-count splitting.",
    )
    parser.add_argument(
        "--max-companies-per-batch",
        type=int,
        default=DEFAULT_MAX_COMPANIES_PER_BATCH,
        help=(
            "Split LLM batches by company count. Default keeps the current registry at regional "
            "batches of up to 6 companies; use 1 for the older one-company fallback."
        ),
    )
    parser.add_argument(
        "--max-estimated-input-tokens",
        type=int,
        default=DEFAULT_MAX_ESTIMATED_INPUT_TOKENS,
        help="Auto-split a generated batch if prompt plus upload-source text is estimated above this input token budget. Use 0 to disable.",
    )
    args = parser.parse_args()

    registry = json.loads(REGISTRY_PATH.read_text(encoding="utf-8"))
    schema_text = SCHEMA_PATH.read_text(encoding="utf-8")
    prompt_template = PROMPT_TEMPLATE_PATH.read_text(encoding="utf-8")
    company_lookup = {company["key"]: company for company in registry["companies"]}
    provider_profile = provider_profiles[args.provider]
    max_upload_files = (
        args.max_upload_files_per_batch
        if args.max_upload_files_per_batch is not None
        else provider_upload_limit(provider_profile)
    )

    run_id = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    output_dir = (args.output_dir / f"{MODE_CONFIG[args.mode]['folder_prefix']}_{run_id}").resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / "aistudio_json").mkdir(exist_ok=True)

    batches = split_batches(registry["batches"], args.max_companies_per_batch)
    batches = split_batches_by_upload_limit(
        batches,
        company_lookup,
        args.mode,
        args.download_mode,
        max_upload_files,
    )
    batch_manifests: list[dict] = []
    for batch in batches:
        batch_manifests.extend(
            prepare_batch_with_budget(
                batch,
                company_lookup,
                output_dir,
                schema_text,
                prompt_template,
                args.mode,
                args.download_mode,
                args.max_estimated_input_tokens,
                max_upload_files,
                args.provider,
                provider_profile,
            )
        )
    write_package_metadata(output_dir, batch_manifests, args.mode, args.provider, provider_profile)
    write_instructions(output_dir, batch_manifests, args.mode, provider_profile)
    (ROOT / "outputs" / f"latest_aistudio_{args.mode}_package_path.txt").write_text(str(output_dir), encoding="utf-8")
    (ROOT / "outputs" / f"latest_aistudio_{args.mode}_{args.provider}_package_path.txt").write_text(
        str(output_dir),
        encoding="utf-8",
    )
    print(output_dir)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
